import sys
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QPushButton,
                             QVBoxLayout, QHBoxLayout, QTabWidget, QLabel,
                             QLineEdit, QFileDialog, QProgressDialog, QMessageBox,
                             QProgressBar, QCheckBox, QScrollArea)  # 添加 QScrollArea
from PySide6.QtCore import Qt, QSettings  # 添加 QSettings 导入
from PySide6.QtGui import QIcon  # 添加这行导入
import openpyxl
from typing import Tuple
import os
import logging
import re
from datetime import datetime

# 辅助函数定义
def standardize_data(value: str, column_index: int) -> str:
    """标准化数据处理"""
    if not value:
        return ""
    
    # 基础清理：去除所有空白字符
    value = ''.join(value.split())
    
    if column_index == 1:
        # 处理中文数字
        cn_num = {'一': '1', '二': '2', '三': '3', '四': '4', '五': '5',
                 '六': '6', '七': '7', '八': '8', '九': '9', '十': '10',
                 '正': '1'}
        
        # 替换中文数字
        for cn, num in cn_num.items():
            value = value.replace(cn, num)
            
        # 首先处理标准日期格式（年月日）
        date_patterns = [
            (r'(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})', '%Y%m'),  # 2024/3/4, 2024-03-04 等
            (r'(\d{4})[-/.](\d{1,2})', '%Y%m'),  # 2024/3, 2024-03 等
            (r'(\d{2})年(\d{1,2})月', '%Y%m'),   # 24年3月
        ]
        
        # 处理中文日期范围
        cn_range_patterns = [
            r'(\d{2,4})年(\d{1,2})月[到至和-](\d{1,2})月',  # 25年3月到4月、25年3月-4月
            r'(\d{2,4})年(\d{1,2})[到至和-](\d{1,2})月',    # 25年3到4月、25年3-4月
        ]
        
        # 处理数字日期范围格式 (如 202411-12)
        num_range_pattern = r'(\d{4})(\d{1,2})-(\d{1,2})'
        match = re.search(num_range_pattern, value)
        if match:
            year = match.group(1)
            start_month = int(match.group(2))
            end_month = int(match.group(3))
            if 1 <= start_month <= 12 and 1 <= end_month <= 12:
                # 返回逗号分隔的月份列表
                months = []
                for month in range(start_month, end_month + 1):
                    months.append(f"{year}{str(month).zfill(2)}")
                return ",".join(months)
        
        # 先尝试匹配中文范围
        for pattern in cn_range_patterns:
            match = re.search(pattern, value)
            if match:
                year = match.group(1)
                if len(year) == 2:
                    year = '20' + year
                start_month = int(match.group(2))
                end_month = int(match.group(3))
                if 1 <= start_month <= 12 and 1 <= end_month <= 12:
                    # 修改返回格式，使用逗号分隔的月份列表
                    months = []
                    for month in range(start_month, end_month + 1):
                        months.append(f"{year}{str(month).zfill(2)}")
                    return ",".join(months)
        
        # 移除月份中的中文字符
        value = value.replace('月', '').replace('年', '')
        
        # 移除或注释掉以下代码块
        # 处理范围表示（例如：3-4月、3到4月）
        # range_patterns = [r'(\d{1,2})[-到至](\d{1,2})', r'(\d{4})[-到至](\d{1,2})']
        # for pattern in range_patterns:
        #     match = re.search(pattern, value)
        #     if match:
        #         # 获取范围的两个月份
        #         start_month = match.group(1)
        #         end_month = match.group(2)
        #         # 返回特殊格式，表示这是一个范围
        #         return f"R{start_month}-{end_month}"
                
        # 标准日期模式匹配
        date_patterns = [
            (r'(\d{4})[-/.]?(\d{1,2})', '%Y%m'),  # 2024/4, 2024-04 等
            (r'(\d{2})(\d{2})', '%Y%m'),  # 2404 等
            (r'(\d{1,2})', '%m'),  # 单独的月份数字
        ]
        
        for pattern, _ in date_patterns:
            match = re.match(pattern, value)
            if match:
                try:
                    groups = match.groups()
                    if len(groups) == 2:
                        year, month = groups
                        if len(year) == 2:
                            year = '20' + year
                    else:
                        # 如果只有月份，使用当前年份
                        year = str(datetime.now().year)
                        month = groups[0]
                    
                    # 确保月份在1-12范围内
                    month = int(month)
                    if 1 <= month <= 12:
                        month = str(month).zfill(2)
                        return f"{year}{month}"
                except:
                    pass
    
    elif column_index == 2:
        # 统一全角和半角字符
        value = value.replace('（', '(').replace('）', ')')
        value = value.replace('：', ':').replace('，', ',')
        value = value.replace('"', '"').replace('"', '"')
        value = value.replace('　', '')  # 全角空格直接移除
        
    elif column_index == 3:
        # 统一标点并转大写
        value = value.replace('（', '(').replace('）', ')')
        value = value.replace('，', ',').replace('：', ':')
        value = value.replace('　', '')  # 全角空格直接移除
        value = value.upper()
    
    return value

def get_sheet_data(sheet, row: int) -> Tuple[str, str, str]:
    """获取并标准化数据"""
    return tuple(standardize_data(str(sheet.cell(row=row, column=i).value), i) for i in range(1, 4))

def clear_sheet(sheet) -> None:
    """清空指定工作表的数据"""
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row)

def copy_title_row(source_sheet, target_sheet) -> None:
    """复制标题行到目标工作表"""
    for column, cell in enumerate(source_sheet[1], start=1):
        target_sheet.cell(row=1, column=column, value=cell.value)

def init_result_sheet(workbook, sheet_name: str):
    """初始化结果工作表"""
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        clear_sheet(sheet)
    else:
        sheet = workbook.create_sheet(sheet_name)
    return sheet

# 修改日志设置部分
settings = QSettings('供应商数据智能匹配系统', 'DataAnalysis')
enable_logging = settings.value('enable_logging', False, bool)

# 创建日志目录
log_dir = os.path.join(os.path.dirname(__file__), 'logs')
if not os.path.exists(log_dir):
    os.makedirs(log_dir)

# 清理历史日志文件
today = datetime.now().strftime("%Y%m%d")
try:
    for log_file_name in os.listdir(log_dir):
        if log_file_name.startswith('供应商匹配_') and log_file_name.endswith('.log'):
            file_date = log_file_name.replace('供应商匹配_', '').replace('.log', '')
            if file_date <= today:
                old_log_path = os.path.join(log_dir, log_file_name)
                try:
                    os.remove(old_log_path)
                    print(f"已清理历史日志: {log_file_name}")
                except Exception as e:
                    print(f"清理日志文件失败 {log_file_name}: {str(e)}")
except Exception as e:
    print(f"清理历史日志时出错: {str(e)}")

# 创建新的日志文件
log_file = os.path.join(log_dir, f'供应商匹配_{today}.log')

if os.path.exists(log_file):
    try:
        os.remove(log_file)
    except Exception as e:
        print(f"清理旧日志文件失败: {str(e)}")

logging.basicConfig(
    filename=log_file,
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# 根据保存的设置决定是否禁用日志
logging.getLogger().disabled = not enable_logging

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.settings = QSettings('供应商数据智能匹配系统', 'DataAnalysis')
        logging.info("初始化供应商数据智能匹配系统")
        self.setWindowTitle("供应商数据智能匹配系统")
        
        # 修改图标设置部分
        icon_path = os.path.join(os.path.dirname(__file__), 'resources', 'icon.ico')
        if getattr(sys, 'frozen', False):
            # 如果是打包后的程序
            base_path = sys._MEIPASS
            icon_path = os.path.join(base_path, 'resources', 'icon.ico')
        
        if os.path.exists(icon_path):
            app_icon = QIcon(icon_path)
            self.setWindowIcon(app_icon)
            QApplication.setWindowIcon(app_icon)  # 设置应用程序级别的图标
            self.setMinimumSize(600, 400)

        # 创建中央部件和主布局
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # 创建标签页
        tab_widget = QTabWidget()
        main_layout.addWidget(tab_widget)

        # 数据筛选标签页
        filter_tab = QWidget()
        filter_layout = QVBoxLayout(filter_tab)

        # 添加帮助文本到筛选页面顶部
        help_widget = QWidget()
        help_layout = QVBoxLayout(help_widget)
        help_toggle = QPushButton("显示/隐藏使用说明")
        help_toggle.setCheckable(True)
        help_toggle.setChecked(False)
        help_layout.addWidget(help_toggle)

        # 创建滚动区域
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll_area.setStyleSheet("""
            QScrollArea {
                border: none;
                background-color: transparent;
            }
            QScrollBar:vertical {
                border: none;
                background: #f0f0f0;
                width: 10px;
                margin: 0px;
            }
            QScrollBar::handle:vertical {
                background: #c0c0c0;
                min-height: 30px;
                border-radius: 5px;
            }
            QScrollBar::handle:vertical:hover {
                background: #a0a0a0;
            }
        """)

        help_content = QLabel('''使用说明：
1. 数据准备：
   - 第一个工作表为"供应商待匹配表"，放入需要查询的数据
   - 第二个工作表为"供应商匹配原表"，放入用于匹配的数据
   - 两个工作表的前三列必须包含：日期、供应商名称、产品名称
   - 请确保Excel文件中只包含这两个工作表，避免干扰分析结果

2. 数据格式要求：
   - 日期格式支持：2024-03、24年3月、3月、202411-12（会自动处理为多个月份）
     示例：2024-03、24年3月、3-4月（会自动处理为多个月份）
   - 供应商名称：不区分全角半角，自动处理空格
     示例："ABC公司"与"A B C公司"会被视为相同
   - 产品名称：不区分大小写，自动处理特殊符号
     示例："Model-A"与"model a"会被视为相同

3. 操作步骤：
   1) 点击"浏览文件"选择Excel文件
   2) 确认数据格式无误后点击"开始分析"
   3) 等待分析完成，查看结果
   4) 分析完成后，结果将保存在同一Excel文件中

4. 处理结果说明：
   - 🟩绿色：表示在匹配原表中找到对应数据
   - 🟥红色：表示在匹配原表中未找到对应数据
   - 🟨黄色：表示该数据重复查询（最高优先级）
   - 🟫棕色：表示日期范围内的数据未能全部匹配成功
   - 🟪紫色：表示日期范围内的数据全部匹配成功
   - 🟧橙色: 表示在表一中重复出现的数据（第二次及以后出现）
   
   颜色优先级：黄色 > 橙色 > 紫色/棕色 > 绿色/红色
   当一条数据符合多个条件时，将按照优先级显示颜色。
   处理逻辑:
   - 系统首先对数据进行标准化处理，统一日期格式、供应商名称和产品名称
   - 对于普通数据，直接在匹配原表中查找对应记录
   - 对于日期范围（如"3-4月"、"202411-12"），系统会检查范围内每个月份是否都能匹配
   - 当一个数据项匹配到多个供应商时，系统会为每个供应商创建单独的记录
   - 匹配结果将分别保存在"匹配到的数据"和"未找到的数据"两个工作表中

5. 常见问题：
   - 如果数据未匹配，请检查日期格式是否正确
   - 供应商名称中的空格和符号会被自动处理
   - 如果分析过程中出现错误，可以在设置中开启日志记录以便排查
   - 大量数据分析可能需要较长时间，请耐心等待''')
        help_content.setWordWrap(True)
        help_content.setStyleSheet("""
            QLabel {
                color: #333333;
                font-size: 13px;
                padding: 10px;
                background-color: #f8f8f8;
                border: 1px solid #e0e0e0;
                border-radius: 4px;
                margin-bottom: 10px;
            }
        """)

        # 将帮助文本放入滚动区域
        scroll_area.setWidget(help_content)
        scroll_area.setVisible(False)  # 初始状态为隐藏
        help_layout.addWidget(scroll_area)

        # 修改显示/隐藏连接
        help_toggle.clicked.connect(lambda checked: scroll_area.setVisible(checked))
        filter_layout.addWidget(help_widget)

        # Excel文件选择部分
        file_widget = QWidget()
        file_layout = QHBoxLayout(file_widget)
        file_layout.setContentsMargins(0, 0, 0, 0)

        file_label = QLabel("选择数据excel")
        self.file_input = QLineEdit()
        self.file_input.setReadOnly(True)
        browse_button = QPushButton("浏览文件")
        browse_button.clicked.connect(self.browse_file)

        file_layout.addWidget(file_label)
        file_layout.addWidget(self.file_input)
        file_layout.addWidget(browse_button)

        filter_layout.addWidget(file_widget)

        # 添加进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        filter_layout.addWidget(self.progress_bar)

        # 开始分析按钮
        analyze_button = QPushButton("开始分析")
        analyze_button.clicked.connect(self.start_analysis)
        analyze_button.setStyleSheet(
            "QPushButton { background-color: #4CAF50; color: white; padding: 8px 16px; border-radius: 4px; }"
            "QPushButton:hover { background-color: #45a049; }"
        )
        filter_layout.addWidget(analyze_button)
        filter_layout.addStretch()

        # 添加标签页
        tab_widget.addTab(filter_tab, "数据筛选")

        # 添加设置页面
        settings_tab = QWidget()
        settings_layout = QVBoxLayout(settings_tab)

        # 添加日志记录选项
        log_checkbox = QCheckBox("启用日志记录")
        log_checkbox.setChecked(self.settings.value('enable_logging', False, bool))
        log_checkbox.stateChanged.connect(self.toggle_logging)
        settings_layout.addWidget(log_checkbox)

        # 添加日志文件位置显示
        log_path_label = QLabel(f"日志文件位置：{os.path.abspath(log_file)}")
        log_path_label.setWordWrap(True)
        settings_layout.addWidget(log_path_label)

        settings_layout.addStretch()
        tab_widget.addTab(settings_tab, "设置")

        # 底部按钮
        button_layout = QHBoxLayout()
        exit_button = QPushButton("退出")
        exit_button.clicked.connect(self.close)
        button_layout.addStretch()
        button_layout.addWidget(exit_button)
        main_layout.addLayout(button_layout)

    def browse_file(self):
        logging.info("开始选择文件")
        try:
            # 从设置中获取上次的目录，如果没有则使用默认目录
            last_dir = self.settings.value('last_directory', os.path.expanduser("~/Documents"))
            
            file_name, _ = QFileDialog.getOpenFileName(
                self,
                "选择Excel文件",
                last_dir,
                "Excel Files (*.xlsx);;All Files (*.*)"
            )
            
            if file_name:
                if not os.path.exists(file_name):
                    logging.error(f"文件不存在: {file_name}")
                    QMessageBox.critical(self, "错误", "所选文件不存在")
                    return
                    
                if not os.access(file_name, os.R_OK):
                    logging.error(f"文件无法访问: {file_name}")
                    QMessageBox.critical(self, "错误", "无法访问所选文件")
                    return
                
                # 保存当前文件的目录路径
                self.settings.setValue('last_directory', os.path.dirname(file_name))
                logging.info(f"选择的文件: {file_name}")
                self.file_input.setText(file_name)
            else:
                logging.info("未选择文件")
                
        except Exception as e:
            logging.error(f"选择文件时出错: {str(e)}")
            QMessageBox.critical(self, "错误", f"选择文件时出错：{str(e)}")

    def toggle_logging(self, state):
        """切换日志状态"""
        self.settings.setValue('enable_logging', bool(state))
        if not state:
            logging.getLogger().disabled = True
            logging.info("日志记录已禁用")
        else:
            logging.getLogger().disabled = False
            logging.info("日志记录已启用")

    def start_analysis(self):
        logging.info("开始数据分析")
        try:
            selected_file = self.file_input.text()
            if not selected_file:
                logging.warning("未选择Excel文件")
                QMessageBox.warning(self, "警告", "请选择Excel文件")
                return
    
            logging.info(f"打开工作簿: {selected_file}")
            workbook = openpyxl.load_workbook(selected_file)
            
            logging.info(f"工作簿包含的工作表: {workbook.sheetnames}")
    
            # 检查工作表数量是否足够
            if len(workbook.worksheets) < 2:
                logging.error("工作簿中缺少必要的工作表")
                QMessageBox.critical(self, "错误", "工作簿中缺少必要的工作表")
                return
    
            # 使用工作表索引而不是名称
            sheet1 = workbook.worksheets[0]  # 第一个工作表
            sheet2 = workbook.worksheets[1]  # 第二个工作表
            sheet3 = init_result_sheet(workbook, "匹配到的数据")
            sheet4 = init_result_sheet(workbook, "未找到的数据")
    
            try:
                self.process_data(workbook, sheet1, sheet2, sheet3, sheet4)
                logging.info("process_data方法执行完成")
            except Exception as e:
                logging.error(f"process_data方法执行失败: {str(e)}")
                logging.error(f"错误类型: {type(e).__name__}")
                logging.error("错误堆栈: ", exc_info=True)
                raise
    
            logging.info(f"准备保存工作簿到: {selected_file}")
            workbook.save(selected_file)
            logging.info("数据分析完成")
            QMessageBox.information(self, "成功", "分析完成")
    
        except Exception as e:
            logging.error(f"分析过程出错: {str(e)}")
            logging.error(f"错误类型: {type(e).__name__}")
            logging.error("错误堆栈: ", exc_info=True)
            QMessageBox.critical(self, "错误", f"执行分析时出错：{str(e)}")

    def process_data(self, workbook, sheet1, sheet2, sheet3, sheet4):
        try:
            logging.info("开始处理数据")
            
            # 复制标题行并添加供应商列
            copy_title_row(sheet1, sheet3)
            copy_title_row(sheet1, sheet4)
            sheet3.cell(row=1, column=4, value="供应商")
            sheet4.cell(row=1, column=4, value="供应商")
    
            # 获取Sheet1的最大行数
            max_row = sheet1.max_row
            if max_row <= 1:
                raise ValueError("Sheet1中没有数据需要匹配")
    
            # 设置进度条
            self.progress_bar.setVisible(True)
            self.progress_bar.setMaximum(max_row - 1)
    
            # 将Sheet2的数据预处理成字典以提高查找效率
            sheet2_data = {}
            for row in sheet2.iter_rows(min_row=2, values_only=True):
                # 使用日期、客户公司、产品名称作为键
                key = (standardize_data(str(row[0]), 1),  # 日期
                      standardize_data(str(row[1]), 2),  # 客户公司
                      standardize_data(str(row[2]), 3))  # 产品名称
                
                # 如果键已存在，则将新的供应商添加到列表中
                if key in sheet2_data:
                    sheet2_data[key].append(row[3])
                else:
                    sheet2_data[key] = [row[3]]  # 创建新列表
    
            # 创建进度对话框
            progress = QProgressDialog("努力分析中....", "取消", 0, max_row - 1, self)
            progress.setWindowModality(Qt.WindowModal)
            progress.setWindowTitle("进度")
    
            # 用于记录已处理的键值
            processed_keys = set()
    
            # 用于记录已添加到匹配表的数据，避免重复
            matched_records = set()
    
            # 用于记录日期范围内的月份
            date_range_map = {}
            
            # 用于记录表一中已经出现过的数据
            sheet1_seen_data = set()

            # 处理每一行数据
            for row in range(2, max_row + 1):
                if progress.wasCanceled():
                    logging.info("用户取消了操作")
                    raise InterruptedError("用户取消了操作")
    
                self.progress_bar.setValue(row - 2)
                logging.debug(f"处理第 {row} 行")
                # 获取原始数据
                original_data = tuple(str(sheet1.cell(row=row, column=i).value) for i in range(1, 4))
                search_key = get_sheet_data(sheet1, row)
                
                logging.debug(f"标准化后的搜索键: {search_key}")
    
                # 初始化状态标记
                is_duplicate = False  # 重复查询
                is_sheet1_duplicate = False  # 表一中重复数据
                is_date_range = False  # 日期范围
                is_date_range_all_match = False  # 日期范围全部匹配
                is_match = False  # 单条数据匹配成功
                matched_results = []  # 存储匹配结果
                
                # 检查当前键是否已处理
                if search_key in processed_keys:
                    is_duplicate = True
                
                # 检查单月是否在已处理的日期范围内
                if not is_duplicate and ',' not in search_key[0]:
                    # 这是单月数据，检查是否包含在已处理的日期范围内
                    for range_key, months in date_range_map.items():
                        if search_key[1:] == range_key and search_key[0] in months:
                            is_duplicate = True
                            break

                # 检查是否为表一中的重复数据
                is_sheet1_duplicate = original_data in sheet1_seen_data
                
                # 将当前数据添加到已见过的数据集合中
                sheet1_seen_data.add(original_data)
                
                # 处理日期范围
                if ',' in search_key[0]:
                    is_date_range = True
                    dates = search_key[0].split(',')
                    all_matches = True
                    
                    # 记录这个日期范围包含的月份
                    date_range_map[search_key[1:]] = dates
                    
                    # 检查范围内的所有日期是否都能匹配
                    for date in dates:
                        test_key = (date,) + search_key[1:]
                        logging.debug(f"检查日期: {date}, 测试键: {test_key}")
                        if test_key not in sheet2_data:
                            all_matches = False
                            logging.debug(f"未匹配的日期: {date}")
                            break
                        # 将每个匹配的结果添加到列表中
                        for supplier in sheet2_data[test_key]:
                            matched_results.append((date, supplier))
                    
                    is_date_range_all_match = all_matches
                
                # 检查单条数据是否匹配
                elif search_key in sheet2_data:
                    is_match = True
                    for supplier in sheet2_data[search_key]:
                        matched_results.append((search_key[0], supplier))
                
                # 根据优先级应用颜色
                fill_color = None
                font_color = '000000'  # 默认黑色
                
                # 颜色优先级：黄色(重复查询) > 橙色(表一重复) > 紫色/棕色(日期范围) > 绿色/红色(单条匹配)
                if is_duplicate:
                    # 黄色 - 重复数据（最高优先级）
                    fill_color = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                elif is_sheet1_duplicate:
                    # 橙色 - 表一中的重复数据（第二优先级）
                    fill_color = openpyxl.styles.PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
                elif is_date_range:
                    if is_date_range_all_match:
                        # 紫色背景，白色字体 - 范围匹配成功
                        fill_color = openpyxl.styles.PatternFill(start_color='9370DB', end_color='9370DB', fill_type='solid')
                        font_color = 'FFFFFF'
                        # 为每个匹配的供应商添加一行，但避免重复
                        for _, supplier in matched_results:
                            # 创建一个唯一标识符，包含公司名称、产品名称和供应商
                            record_key = (search_key[1], search_key[2], supplier)
                            if record_key not in matched_records:
                                sheet3.append(original_data + (supplier,))
                                matched_records.add(record_key)
                    else:
                        # 棕色背景，白色字体 - 范围匹配失败
                        fill_color = openpyxl.styles.PatternFill(start_color='8B4513', end_color='8B4513', fill_type='solid')
                        font_color = 'FFFFFF'
                        sheet4.append(original_data + ('',))
                elif is_match:
                    # 绿色 - 匹配成功
                    fill_color = openpyxl.styles.PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                    # 为每个匹配的供应商添加一行，但避免重复
                    for _, supplier in matched_results:
                        # 创建一个唯一标识符，包含公司名称、产品名称和供应商
                        record_key = (search_key[1], search_key[2], supplier)
                        if record_key not in matched_records:
                            sheet3.append(original_data + (supplier,))
                            matched_records.add(record_key)
                else:
                    # 浅红色 - 未找到匹配
                    fill_color = openpyxl.styles.PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                    sheet4.append(original_data + ('',))
    
                # 应用颜色到原始数据行
                for col in range(1, 4):
                    cell = sheet1.cell(row=row, column=col)
                    cell.fill = fill_color
                    cell.font = openpyxl.styles.Font(color=font_color)

                # 将当前键添加到已处理集合中
                processed_keys.add(search_key)

            progress.setValue(max_row - 1)
            self.progress_bar.setVisible(False)
                
        except Exception as e:
            self.progress_bar.setVisible(False)
            logging.error(f"数据处理出错: {str(e)}")
            logging.error(f"错误类型: {type(e).__name__}")
            logging.error(f"错误堆栈: ", exc_info=True)
            raise

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())


