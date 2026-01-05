"""
主窗口模块

本模块定义了供应商数据智能匹配系统的主窗口类。
负责应用程序的整体布局、菜单栏、标签页管理和数据处理流程的协调。

主要功能:
- 创建和管理应用程序主窗口
- 协调标签页组件(数据筛选页、设置页)
- 处理文件选择和数据分析流程
- 管理数据匹配算法和结果展示
- 处理菜单栏和帮助系统
- 线程管理和进度更新

依赖模块:
- ui.tabs.filter_tab: 数据筛选标签页
- ui.tabs.settings_tab: 设置标签页
- ui.widgets.help_widget: 帮助组件
- core.*: 核心数据处理模块

作者: 供应商数据智能匹配系统开发团队
版本: 1.0
"""

import sys
import os
import logging
from typing import Dict, Set, Tuple, List, Optional

from PySide6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QTabWidget,
    QFileDialog, QMessageBox, QPushButton, QHBoxLayout
)
from PySide6.QtCore import QSettings, Qt
from PySide6.QtGui import QIcon
import openpyxl

# 导入标签页组件
from ui.tabs.filter_tab import FilterTab
from ui.tabs.settings_tab import SettingsTab

# 导入帮助组件
from ui.widgets.help_widget import HelpWidget

# 导入核心模块
from core.data_models import MatchResult, CellStyles
from core.data_standardizer import standardize_data
from core.excel_processor import (
    get_sheet_data, clear_sheet, copy_title_row, init_result_sheet
)
from core.logging_config import setup_logging


class MainWindow(QMainWindow):
    """
    供应商数据智能匹配系统主窗口

    这是应用程序的主窗口类，负责协调所有组件的工作。
    提供数据筛选、分析、统计和设置等功能。

    主要职责:
        1. 窗口初始化和UI布局
        2. 标签页管理(筛选页、设置页)
        3. 菜单栏创建(包含帮助菜单)
        4. 文件选择和验证
        5. 数据分析流程控制
        6. 数据匹配算法实现
        7. 统计信息更新
        8. 日志系统管理

    属性:
        settings (QSettings): 应用程序配置对象
        log_file (str): 日志文件路径
        recent_files (List[str]): 最近打开的文件列表
        filter_tab (FilterTab): 数据筛选标签页
        settings_tab (SettingsTab): 设置标签页

    示例:
        >>> app = QApplication(sys.argv)
        >>> window = MainWindow()
        >>> window.show()
        >>> sys.exit(app.exec())
    """

    def __init__(self):
        """
        初始化主窗口

        执行以下初始化步骤:
        1. 调用父类初始化
        2. 创建配置对象
        3. 初始化日志系统
        4. 创建用户界面
        5. 记录初始化完成日志
        """
        super().__init__()

        # 应用程序配置
        self.settings = QSettings('供应商数据智能匹配系统', 'DataAnalysis')
        self.log_file: Optional[str] = None
        self.recent_files: List[str] = []

        # 初始化系统
        self._init_logging()
        self._init_ui()

        logging.info("系统初始化完成")

    def _init_logging(self):
        """
        初始化日志系统

        创建日志目录，配置日志格式，并根据设置启用或禁用日志记录。
        """
        # 获取脚本所在目录
        if getattr(sys, 'frozen', False):
            # 打包后的可执行文件
            base_dir = sys._MEIPASS
        else:
            # 开发环境
            base_dir = os.path.dirname(os.path.abspath(__file__))

        log_dir = os.path.join(os.path.dirname(base_dir), 'logs')

        # 设置日志系统
        self.log_file = setup_logging(log_dir)

        # 根据配置启用或禁用日志
        enable_logging = self.settings.value('enable_logging', False, bool)
        logging.getLogger().disabled = not enable_logging

    def _init_ui(self):
        """
        初始化用户界面

        创建主窗口的所有UI组件，包括:
        - 窗口标题和图标
        - 中央部件和布局
        - 标签页组件
        - 菜单栏
        - 底部按钮
        """
        # 设置窗口基本属性
        self.setWindowTitle("供应商数据智能匹配系统")
        self._set_window_icon()
        self.setMinimumSize(800, 600)

        # 设置应用样式
        self._set_app_style()

        # 创建中央部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # 创建标签页组件
        tab_widget = self._create_tab_widget()
        main_layout.addWidget(tab_widget)

        # 创建底部按钮
        self._create_bottom_buttons(main_layout)

        # 创建菜单栏
        self._create_menu_bar()

        # 创建状态栏
        self._create_status_bar()

    def _set_window_icon(self):
        """
        设置窗口图标

        从资源目录加载图标文件，并设置为窗口图标。
        支持开发环境和打包环境两种路径。
        """
        # 开发环境路径
        icon_path = os.path.join(
            os.path.dirname(os.path.dirname(__file__)),
            'resources', 'icon.ico'
        )

        # 打包环境路径
        if getattr(sys, 'frozen', False):
            base_path = sys._MEIPASS
            icon_path = os.path.join(base_path, 'resources', 'icon.ico')

        # 设置图标
        if os.path.exists(icon_path):
            app_icon = QIcon(icon_path)
            self.setWindowIcon(app_icon)
            from PySide6.QtWidgets import QApplication
            QApplication.setWindowIcon(app_icon)

    def _set_app_style(self):
        """
        设置应用程序样式

        配置应用程序的全局样式，包括:
        - 字体和颜色
        - 滚动条样式
        - 按钮样式
        - 输入框样式
        - 进度条样式
        """
        from PySide6.QtWidgets import QApplication

        # 使用Fusion风格
        QApplication.setStyle("Fusion")

        # 设置全局样式表
        self.setStyleSheet("""
            QMainWindow {
                background-color: #F5F5F5;
            }
            QWidget {
                font-family: "Microsoft YaHei", "PingFang SC", sans-serif;
                font-size: 13px;
                color: #546E7A;
            }
            /* 苹果风格滚动条 - 垂直 */
            QScrollBar:vertical {
                border: none;
                background: transparent;
                width: 10px;
                margin: 0px;
            }
            QScrollBar::handle:vertical {
                background: #C1C1C1;
                min-height: 30px;
                border-radius: 5px;
                margin: 2px;
            }
            QScrollBar::handle:vertical:hover {
                background: #A8A8A8;
            }
            QScrollBar::handle:vertical:pressed {
                background: #8F8F8F;
            }
            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {
                border: none;
                background: none;
                height: 0px;
            }
            QScrollBar::add-page:vertical,
            QScrollBar::sub-page:vertical {
                background: none;
            }
            /* 苹果风格滚动条 - 水平 */
            QScrollBar:horizontal {
                border: none;
                background: transparent;
                height: 10px;
                margin: 0px;
            }
            QScrollBar::handle:horizontal {
                background: #C1C1C1;
                min-width: 30px;
                border-radius: 5px;
                margin: 2px;
            }
            QScrollBar::handle:horizontal:hover {
                background: #A8A8A8;
            }
            QScrollBar::handle:horizontal:pressed {
                background: #8F8F8F;
            }
            QScrollBar::add-line:horizontal,
            QScrollBar::sub-line:horizontal {
                border: none;
                background: none;
                width: 0px;
            }
            QScrollBar::add-page:horizontal,
            QScrollBar::sub-page:horizontal {
                background: none;
            }
            QScrollArea {
                border: none;
                background-color: transparent;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 6px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #3d8b40;
            }
            QPushButton:disabled {
                background-color: #E0E0E0;
                color: #9E9E9E;
            }
            QLineEdit {
                padding: 8px 12px;
                border: 1px solid #CFD8DC;
                border-radius: 4px;
                background-color: white;
                font-size: 13px;
                color: #546E7A;
            }
            QLineEdit:focus {
                border: 1px solid #4CAF50;
            }
            QProgressBar {
                border: none;
                border-radius: 4px;
                background-color: #E0E0E0;
                text-align: center;
                font-size: 12px;
                color: #546E7A;
            }
            QProgressBar::chunk {
                background-color: #4CAF50;
                border-radius: 4px;
            }
            QGroupBox {
                border: 1px solid #E0E0E0;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
                font-weight: bold;
                background-color: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: #607D8B;
            }
        """)

    def _create_tab_widget(self) -> QTabWidget:
        """
        创建标签页组件

        创建并配置标签页组件，添加数据筛选和设置两个标签页。

        返回:
            QTabWidget: 配置好的标签页组件
        """
        # 创建标签页
        tab_widget = QTabWidget()

        # 设置标签页样式
        tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: none;
                background-color: transparent;
            }
            QTabBar::tab {
                background-color: #F5F5F5;
                color: #607D8B;
                padding: 10px 20px;
                margin-right: 5px;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
                font-size: 13px;
            }
            QTabBar::tab:selected {
                background-color: white;
                color: #4CAF50;
                font-weight: bold;
            }
            QTabBar::tab:hover:!selected {
                background-color: #EEEEEE;
                color: #546E7A;
            }
        """)

        # 创建数据筛选标签页
        self.filter_tab = FilterTab()
        self.filter_tab.file_browsed.connect(self._on_file_browsed)
        self.filter_tab.analyze_clicked.connect(self._on_analyze_clicked)
        self.filter_tab.file_dropped.connect(self._on_file_dropped)
        tab_widget.addTab(self.filter_tab, "📊 数据筛选")

        # 创建设置标签页
        self.settings_tab = SettingsTab()
        self.settings_tab.set_log_file(self.log_file)
        self.settings_tab.logging_toggled.connect(self._on_logging_toggled)
        tab_widget.addTab(self.settings_tab, "⚙️ 设置")

        return tab_widget

    def _create_menu_bar(self):
        """
        创建菜单栏

        添加应用程序的菜单栏，包括帮助菜单。
        """
        from PySide6.QtWidgets import QMenuBar, QMenu

        menubar = self.menuBar()

        # 创建帮助菜单
        help_menu = QMenu("帮助", self)
        menubar.addMenu(help_menu)

        # 添加使用说明
        help_action = help_menu.addAction("使用说明")
        help_action.triggered.connect(self._show_help_dialog)

        # 添加关于
        about_action = help_menu.addAction("关于")
        about_action.triggered.connect(self._show_about_dialog)

    def _create_status_bar(self):
        """
        创建状态栏

        添加应用程序的状态栏，用于显示提示信息。
        """
        from PySide6.QtWidgets import QStatusBar

        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("就绪")

    def _create_bottom_buttons(self, layout: QVBoxLayout):
        """
        创建底部按钮

        在主布局底部添加退出按钮。

        参数:
            layout: 主窗口的垂直布局
        """
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)
        button_layout.addStretch()

        # 退出按钮
        exit_button = QPushButton("退出")
        exit_button.setStyleSheet("""
            QPushButton {
                background-color: #F5F5F5;
                color: #666666;
                border: 1px solid #DDDDDD;
                padding: 8px 20px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #EEEEEE;
            }
        """)
        exit_button.clicked.connect(self.close)
        button_layout.addWidget(exit_button)

        layout.addLayout(button_layout)

    # ==================== 事件处理函数 ====================

    def _on_file_browsed(self):
        """
        处理浏览文件按钮点击事件

        打开文件选择对话框，让用户选择Excel文件。
        """
        logging.info("用户点击浏览文件按钮")
        self._browse_file()

    def _on_analyze_clicked(self):
        """
        处理开始分析按钮点击事件

        启动数据分析流程。
        """
        logging.info("用户点击开始分析按钮")
        self._start_analysis()

    def _on_file_dropped(self, file_path: str):
        """
        处理文件拖放事件

        当用户拖放文件到筛选标签页时调用。

        参数:
            file_path: 拖放的文件路径
        """
        if os.path.exists(file_path):
            self.filter_tab.set_file_path(file_path)
            logging.info(f"通过拖拽选择文件: {file_path}")
            self.status_bar.showMessage(f"已选择文件: {os.path.basename(file_path)}")
        else:
            QMessageBox.warning(self, "警告", "文件不存在")
            logging.error(f"文件不存在: {file_path}")

    def _on_logging_toggled(self, enabled: bool):
        """
        处理日志设置切换事件

        当用户在设置页切换日志开关时调用。

        参数:
            enabled: 日志是否启用
        """
        status = "启用" if enabled else "禁用"
        self.status_bar.showMessage(f"日志记录已{status}")

    # ==================== 文件处理函数 ====================

    def _browse_file(self):
        """
        浏览并选择Excel文件

        打开文件选择对话框，验证选择的文件，并更新界面。
        """
        logging.info("开始选择文件")

        try:
            # 获取上次打开的目录
            last_dir = self.settings.value(
                'last_directory',
                os.path.expanduser("~/Documents")
            )

            # 打开文件选择对话框
            file_name, _ = QFileDialog.getOpenFileName(
                self,
                "选择Excel文件",
                last_dir,
                "Excel Files (*.xlsx *.xls);;All Files (*.*)"
            )

            if file_name:
                self._validate_and_set_file(file_name)
            else:
                logging.info("未选择文件")

        except Exception as e:
            logging.error(f"选择文件时出错: {str(e)}", exc_info=True)
            QMessageBox.critical(self, "错误", f"选择文件时出错：{str(e)}")

    def _validate_and_set_file(self, file_name: str):
        """
        验证并设置文件路径

        检查文件是否存在和可访问，然后更新界面和配置。

        参数:
            file_name: 文件路径
        """
        # 检查文件是否存在
        if not os.path.exists(file_name):
            logging.error(f"文件不存在: {file_name}")
            QMessageBox.critical(self, "错误", "所选文件不存在")
            return

        # 检查文件是否可读
        if not os.access(file_name, os.R_OK):
            logging.error(f"文件无法访问: {file_name}")
            QMessageBox.critical(self, "错误", "无法访问所选文件")
            return

        # 保存配置和更新界面
        self.settings.setValue('last_directory', os.path.dirname(file_name))
        self.filter_tab.set_file_path(file_name)

        logging.info(f"选择的文件: {file_name}")
        self.status_bar.showMessage(f"已选择文件: {os.path.basename(file_name)}")

    # ==================== 数据分析函数 ====================

    def _start_analysis(self):
        """
        开始数据分析

        执行完整的数据分析流程:
        1. 验证文件选择
        2. 加载Excel文件
        3. 处理数据匹配
        4. 保存结果
        5. 更新统计信息
        """
        # 检查是否选择了文件
        file_path = self.filter_tab.get_file_path()
        if not file_path:
            QMessageBox.warning(self, "警告", "请先选择Excel文件")
            return

        logging.info("开始数据分析")

        # 禁用分析按钮
        self.filter_tab.enable_analyze_button(False)

        try:
            # 加载工作簿
            workbook = openpyxl.load_workbook(file_path)
            logging.info(f"工作簿包含的工作表: {workbook.sheetnames}")

            # 检查工作表数量
            if len(workbook.worksheets) < 2:
                QMessageBox.critical(self, "错误", "工作簿中缺少必要的工作表")
                self.filter_tab.enable_analyze_button(True)
                return

            # 获取工作表
            sheet1 = workbook.worksheets[0]  # 待匹配表
            sheet2 = workbook.worksheets[1]  # 匹配原表
            sheet3 = init_result_sheet(workbook, "匹配到的数据")
            sheet4 = init_result_sheet(workbook, "未找到的数据")

            # 显示进度条
            self.filter_tab.set_progress_visible(True)

            # 处理数据
            stats = self._process_data(workbook, sheet1, sheet2, sheet3, sheet4)

            # 保存结果
            workbook.save(file_path)
            self.filter_tab.set_progress_visible(False)

            # 更新统计信息
            self.filter_tab.update_stats(stats)

            # 显示完成消息
            logging.info("数据分析完成")
            QMessageBox.information(
                self,
                "✅ 分析完成",
                f"数据处理完成！\n\n"
                f"总计：{stats['total']} 条\n"
                f"已匹配：{stats['matched']} 条\n"
                f"未匹配：{stats['unmatched']} 条\n"
                f"匹配率：{stats['rate']}%",
                QMessageBox.Ok
            )

            self.status_bar.showMessage("分析完成")

        except Exception as e:
            self.filter_tab.set_progress_visible(False)
            logging.error(f"分析过程出错: {str(e)}", exc_info=True)
            QMessageBox.critical(self, "错误", f"执行分析时出错：{str(e)}")
            self.status_bar.showMessage("分析失败")

        finally:
            # 重新启用分析按钮
            self.filter_tab.enable_analyze_button(True)

    def _process_data(self, workbook, sheet1, sheet2, sheet3, sheet4) -> Dict[str, any]:
        """
        处理数据匹配逻辑

        执行数据匹配的核心算法，包括:
        1. 初始化结果表
        2. 预处理匹配数据
        3. 逐行分析匹配
        4. 应用样式标记
        5. 分类结果

        参数:
            workbook: Excel工作簿对象
            sheet1: 待匹配表
            sheet2: 匹配原表
            sheet3: 匹配结果表
            sheet4: 未匹配结果表

        返回:
            包含统计信息的字典，包括total, matched, unmatched, rate
        """
        try:
            logging.info("开始处理数据")

            # 初始化结果表
            copy_title_row(sheet1, sheet3)
            copy_title_row(sheet1, sheet4)
            sheet3.cell(row=1, column=4, value="供应商")
            sheet4.cell(row=1, column=4, value="供应商")

            # 预处理匹配数据
            sheet2_data = self._preprocess_sheet2(sheet2)

            # 检查数据量
            max_row = sheet1.max_row
            if max_row <= 1:
                raise ValueError("Sheet1中没有数据需要匹配")

            # 设置进度条
            self.filter_tab.set_progress_maximum(max_row - 1)

            # 处理数据
            processed_keys: Set[Tuple[str, str, str]] = set()
            date_range_map: Dict[Tuple[str, str], List[str]] = {}

            matched_count = 0
            unmatched_count = 0

            for row in range(2, max_row + 1):
                # 更新进度
                self.filter_tab.set_progress_value(row - 2)

                # 获取原始数据和标准化后的搜索键
                original_data = tuple(
                    str(sheet1.cell(row=row, column=i).value)
                    for i in range(1, 4)
                )
                search_key = get_sheet_data(sheet1, row)

                # 分析匹配
                result = self._analyze_match(
                    search_key, sheet2_data, processed_keys, date_range_map
                )

                # 应用样式
                cell_style = self._determine_cell_style(result)
                for col in range(1, 4):
                    cell = sheet1.cell(row=row, column=col)
                    cell.fill = cell_style.to_pattern_fill()
                    cell.font = cell_style.to_font()

                # 保存结果
                if result.is_match or (result.is_date_range and result.is_all_match):
                    matched_count += 1
                    target_sheet = sheet3
                    for _, supplier in result.matched_suppliers:
                        target_sheet.append(original_data + (supplier,))
                else:
                    unmatched_count += 1
                    sheet4.append(original_data + ('',))

                # 标记为已处理
                processed_keys.add(search_key)

            # 计算统计
            total = matched_count + unmatched_count
            rate = (matched_count / total * 100) if total > 0 else 0

            return {
                'total': total,
                'matched': matched_count,
                'unmatched': unmatched_count,
                'rate': f"{rate:.1f}"
            }

        except Exception as e:
            self.filter_tab.set_progress_visible(False)
            logging.error(f"数据处理出错: {str(e)}", exc_info=True)
            raise

    def _preprocess_sheet2(self, sheet2) -> Dict[Tuple[str, str, str], List[str]]:
        """
        预处理匹配原表数据

        将sheet2的数据转换为字典结构，便于快速查找。

        参数:
            sheet2: 匹配原表

        返回:
            字典，键为(日期, 客户名称, 产品名称)，值为供应商列表
        """
        sheet2_data = {}

        for row in sheet2.iter_rows(min_row=2, values_only=True):
            key = (
                standardize_data(str(row[0]), 1),
                standardize_data(str(row[1]), 2),
                standardize_data(str(row[2]), 3)
            )

            if key in sheet2_data:
                sheet2_data[key].append(row[3])
            else:
                sheet2_data[key] = [row[3]]

        return sheet2_data

    def _analyze_match(
        self,
        search_key: Tuple[str, str, str],
        sheet2_data: Dict,
        processed_keys: Set[Tuple[str, str, str]],
        date_range_map: Dict[Tuple[str, str], List[str]]
    ) -> MatchResult:
        """
        分析数据匹配情况

        判断待匹配数据是否在匹配原表中存在，处理重复数据和日期范围数据。

        参数:
            search_key: 标准化后的搜索键(日期, 客户, 产品)
            sheet2_data: 预处理后的匹配原表数据
            processed_keys: 已处理的键集合
            date_range_map: 日期范围映射

        返回:
            MatchResult对象，包含匹配结果信息
        """
        result = MatchResult()

        # 检查是否重复
        result.is_duplicate = self._check_duplicate(
            search_key, processed_keys, date_range_map
        )

        # 日期范围数据处理
        if ',' in search_key[0]:
            result.is_date_range = True
            dates = search_key[0].split(',')
            date_range_map[search_key[1:]] = dates

            all_matches = True
            for date in dates:
                test_key = (date,) + search_key[1:]
                if test_key in sheet2_data:
                    for supplier in sheet2_data[test_key]:
                        result.matched_suppliers.append((date, supplier))
                else:
                    all_matches = False

            result.is_all_match = all_matches and bool(result.matched_suppliers)

        # 单条数据处理
        elif not result.is_duplicate and search_key in sheet2_data:
            result.is_match = True
            for supplier in sheet2_data[search_key]:
                result.matched_suppliers.append((search_key[0], supplier))

        return result

    def _check_duplicate(
        self,
        search_key: Tuple[str, str, str],
        processed_keys: Set[Tuple[str, str, str]],
        date_range_map: Dict[Tuple[str, str], List[str]]
    ) -> bool:
        """
        检查是否为重复数据

        判断当前数据是否在之前已经处理过，包括直接重复和日期范围展开后的重复。

        参数:
            search_key: 待检查的搜索键
            processed_keys: 已处理的键集合
            date_range_map: 日期范围映射

        返回:
            bool: 如果是重复数据返回True，否则返回False
        """
        # 检查是否直接重复
        if search_key in processed_keys:
            return True

        # 检查是否在日期范围内
        if ',' not in search_key[0]:
            for range_key, months in date_range_map.items():
                if search_key[1:] == range_key and search_key[0] in months:
                    return True

        # 如果是日期范围，检查其展开的月份是否重复
        if ',' in search_key[0]:
            dates = search_key[0].split(',')
            for date in dates:
                single_key = (date,) + search_key[1:]
                if single_key in processed_keys:
                    return True

        return False

    def _determine_cell_style(self, result: MatchResult):
        """
        根据匹配结果确定单元格样式

        根据匹配结果的状态选择相应的颜色标记。

        样式优先级:
            1. 黄色 - 重复数据（最高优先级）
            2. 紫色 - 日期范围且全部匹配
            3. 棕色 - 日期范围但未全部匹配
            4. 绿色 - 单条数据匹配成功
            5. 红色 - 单条数据未匹配

        参数:
            result: 匹配结果对象

        返回:
            CellStyle对象
        """
        if result.is_duplicate:
            return CellStyles.YELLOW
        elif result.is_date_range:
            return CellStyles.PURPLE if result.is_all_match else CellStyles.BROWN
        elif result.is_match:
            return CellStyles.GREEN
        else:
            return CellStyles.RED

    # ==================== 帮助和对话框函数 ====================

    def _show_help_dialog(self):
        """
        显示帮助对话框

        使用HelpWidget显示详细的使用说明。
        """
        HelpWidget.show_detailed_help_dialog(self)

    def _show_about_dialog(self):
        """
        显示关于对话框

        显示应用程序的版本信息和特性说明。
        """
        QMessageBox.information(
            self,
            "关于",
            "供应商数据智能匹配系统 v1.0\n\n"
            "一个现代化的Excel数据匹配工具，\n"
            "帮助您快速处理和分析供应商数据。\n\n"
            "特性：\n"
            "• 支持拖拽上传\n"
            "• 智能日期范围处理\n"
            "• 实时统计显示\n"
            "• 简洁现代界面",
            QMessageBox.Ok
        )
