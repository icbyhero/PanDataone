"""
Excel处理模块

本模块提供Excel工作表的基础操作功能，包括数据获取、工作表清理、标题行复制等。
这些函数是数据处理流程的基础组件。

主要功能:
- 从工作表获取并标准化数据
- 清空工作表数据
- 复制标题行到目标工作表
- 初始化结果工作表

依赖:
- openpyxl: 用于读写Excel文件
- data_standardizer: 用于数据标准化

作者: 供应商数据智能匹配系统开发团队
版本: 1.0
"""

from typing import Tuple
import openpyxl
import logging
from .data_standardizer import standardize_data


def get_sheet_data(sheet, row: int) -> Tuple[str, str, str]:
    """
    获取并标准化工作表数据

    从指定工作表的指定行读取前三列数据，并对每列数据进行标准化处理。
    返回包含三个标准化后数据的元组。

    参数:
        sheet: openpyxl的工作表对象
        row (int): 要读取的行号（从1开始）

    返回:
        Tuple[str, str, str]: 包含三个标准化数据的元组
            - 第一个元素: 标准化后的日期
            - 第二个元素: 标准化后的客户名称
            - 第三个元素: 标准化后的产品名称

    处理流程:
        1. 读取指定行的第1、2、3列单元格
        2. 将每个单元格的值转换为字符串
        3. 根据列号调用相应的标准化函数
        4. 记录调试日志
        5. 返回标准化后的数据元组

    错误处理:
        - 如果单元格值为None，转换为字符串"None"后进行标准化
        - 标准化函数会处理空值等异常情况

    示例:
        >>> # 读取第2行数据
        >>> data = get_sheet_data(worksheet, 2)
        >>> print(data)
        ('202403', '客户A', '产品B')

        >>> # 读取第10行数据
        >>> date, customer, product = get_sheet_data(worksheet, 10)

    注意:
        - 行号从1开始，第1行通常是标题行
        - 只读取前三列数据，分别对应日期、客户名称、产品名称
        - 返回的数据已经过标准化，可以直接用于比较
    """
    # 使用列表推导式读取并标准化前三列数据
    values = tuple(
        standardize_data(str(sheet.cell(row=row, column=i).value), i)
        for i in range(1, 4)  # 列索引1, 2, 3
    )

    # 记录调试日志，便于追踪数据处理过程
    logging.debug(f"行{row}原始数据: {values}")

    return values


def clear_sheet(sheet) -> None:
    """
    清空工作表数据

    删除工作表中的所有数据行（保留标题行）。
    通常用于重新写入数据前清理旧数据。

    参数:
        sheet: openpyxl的工作表对象

    返回:
        None

    处理逻辑:
        1. 检查工作表是否有数据行（max_row > 1）
        2. 如果有数据，从第2行开始删除到最后一行
        3. 保留第1行（标题行）不删除

    错误处理:
        - 如果工作表只有标题行或为空，不执行任何操作
        - openpyxl会自动处理工作表边界检查

    示例:
        >>> # 清空工作表数据
        >>> clear_sheet(worksheet)
        >>> # 工作表现在只有标题行

    注意:
        - 此操作不可逆，删除的数据无法恢复
        - 确保在删除前已保存需要的数据
        - 删除操作会直接修改工作表对象
    """
    # 检查是否有数据行需要删除（max_row > 1表示至少有数据行）
    if sheet.max_row > 1:
        # 删除从第2行到最后一行的所有数据
        sheet.delete_rows(2, sheet.max_row)


def copy_title_row(source_sheet, target_sheet) -> None:
    """
    复制标题行

    将源工作表的标题行（第1行）复制到目标工作表的第1行。
    用于创建具有相同标题的新工作表。

    参数:
        source_sheet: 源工作表对象（复制标题行的工作表）
        target_sheet: 目标工作表对象（接收标题行的工作表）

    返回:
        None

    处理逻辑:
        1. 遍历源工作表第1行的所有单元格
        2. 将每个单元格的值复制到目标工作表的对应位置
        3. 只复制单元格的值，不复制格式

    错误处理:
        - 如果源工作表第1行为空，目标工作表第1行将被清空
        - 不会抛出异常，即使源工作表不存在

    示例:
        >>> # 从sheet1复制标题到sheet2
        >>> copy_title_row(sheet1, sheet2)
        >>> # sheet2的第1行现在与sheet1的第1行相同

    使用场景:
        - 创建结果工作表时保留原标题
        - 统一多个工作表的标题格式
        - 初始化新工作表的结构

    注意:
        - 只复制单元格的值，不复制单元格的格式（字体、颜色等）
        - 如果需要复制格式，需要额外处理
        - 目标工作表第1行的原有数据会被覆盖
    """
    # 遍历源工作表第1行的所有单元格
    for column, cell in enumerate(source_sheet[1], start=1):
        # 将单元格值复制到目标工作表的对应位置
        target_sheet.cell(row=1, column=column, value=cell.value)


def init_result_sheet(workbook, sheet_name: str):
    """
    初始化结果工作表

    创建或重置一个用于存储结果的工作表。
    如果工作表已存在，清空其数据；如果不存在，创建新的工作表。

    参数:
        workbook: openpyxl的工作簿对象
        sheet_name (str): 工作表名称

    返回:
        openpyxl的工作表对象: 初始化后的工作表

    处理流程:
        1. 检查工作簿中是否已存在指定名称的工作表
        2. 如果存在：
           - 获取该工作表的引用
           - 清空工作表中的所有数据行（保留标题行）
        3. 如果不存在：
           - 创建新的工作表
           - 返回新创建的工作表对象
        4. 返回工作表对象供后续操作使用

    错误处理:
        - 工作表名称冲突由openpyxl自动处理
        - 如果工作表名称不符合Excel规范，openpyxl会抛出异常

    示例:
        >>> # 创建或重置"匹配到的数据"工作表
        >>> result_sheet = init_result_sheet(workbook, "匹配到的数据")
        >>> # 现在可以向result_sheet写入数据

        >>> # 创建或重置"未找到的数据"工作表
        >>> unmatched_sheet = init_result_sheet(workbook, "未找到的数据")

    使用场景:
        - 在数据分析前准备结果存储工作表
        - 重新运行分析时清空旧结果
        - 创建标准化的结果输出表

    注意:
        - 返回的工作表可能包含标题行（如果是已存在的工作表）
        - 新创建的工作表是空白的，需要手动添加标题和数据
        - 工作表名称会自动添加到工作簿的sheetnames列表中
    """
    # 检查工作簿中是否已存在指定名称的工作表
    if sheet_name in workbook.sheetnames:
        # 工作表已存在，获取其引用并清空数据
        sheet = workbook[sheet_name]
        clear_sheet(sheet)
    else:
        # 工作表不存在，创建新的工作表
        sheet = workbook.create_sheet(sheet_name)

    # 返回工作表对象供后续使用
    return sheet
