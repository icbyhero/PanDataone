"""
数据模型定义模块

本模块定义了供应商数据智能匹配系统中使用的核心数据类。

主要功能:
- 定义匹配结果数据结构 (MatchResult)
- 定义单元格样式配置 (CellStyle)
- 提供预定义的单元格样式集合 (CellStyles)

作者: 供应商数据智能匹配系统开发团队
版本: 1.0
"""

from dataclasses import dataclass
from typing import List, Tuple
from openpyxl.styles import PatternFill, Font


@dataclass
class MatchResult:
    """
    匹配结果数据类

    用于存储数据匹配分析的结果信息，包含多种匹配状态的标识。

    属性:
        is_duplicate (bool): 是否为重复数据。重复数据指在之前的处理中已经出现过相同的数据项。
        is_date_range (bool): 是否为日期范围数据。日期范围数据包含多个连续的月份（如"202401-03"）。
        is_all_match (bool): 日期范围内的所有数据是否全部匹配成功。仅当is_date_range为True时有意义。
        is_match (bool): 是否成功匹配到对应数据。对于单条数据，表示在匹配原表中找到了对应记录。
        matched_suppliers (List[Tuple[str, str]]): 匹配到的供应商列表。
            每个元素是一个元组 (日期, 供应商名称)。
            对于日期范围数据，可能包含多个供应商信息。
            对于单条数据，通常只包含一个或多个供应商。

    示例:
        >>> result = MatchResult()
        >>> result.is_match = True
        >>> result.matched_suppliers = [("202401", "供应商A")]

        >>> date_range_result = MatchResult()
        >>> date_range_result.is_date_range = True
        >>> date_range_result.is_all_match = True
        >>> date_range_result.matched_suppliers = [
        ...     ("202401", "供应商A"),
        ...     ("202402", "供应商B"),
        ...     ("202403", "供应商C")
        ... ]
    """

    is_duplicate: bool = False
    is_date_range: bool = False
    is_all_match: bool = False
    is_match: bool = False
    matched_suppliers: List[Tuple[str, str]] = None

    def __post_init__(self):
        """
        数据类初始化后处理

        确保matched_suppliers列表被正确初始化为空列表，避免可变默认参数的问题。

        说明:
            如果matched_suppliers为None（默认值），则将其初始化为空列表。
            这是Python中处理可变默认参数的最佳实践。
        """
        if self.matched_suppliers is None:
            self.matched_suppliers = []


@dataclass
class CellStyle:
    """
    单元格样式配置类

    用于定义Excel单元格的样式，包括填充颜色和字体颜色。
    提供转换为openpyxl样式对象的方法。

    属性:
        fill_color (str): 填充颜色，使用十六进制颜色码（如'FFFF00'表示黄色）。
        font_color (str): 字体颜色，使用十六进制颜色码，默认为'000000'（黑色）。

    方法:
        to_pattern_fill(): 将样式转换为openpyxl的PatternFill对象
        to_font(): 将样式转换为openpyxl的Font对象

    示例:
        >>> # 创建黄色背景的样式
        >>> yellow_style = CellStyle('FFFF00')
        >>> fill = yellow_style.to_pattern_fill()
        >>> font = yellow_style.to_font()

        >>> # 创建自定义样式
        >>> custom_style = CellStyle('9370DB', 'FFFFFF')
        >>> fill = custom_style.to_pattern_fill()
        >>> font = custom_style.to_font()
    """

    fill_color: str
    font_color: str = '000000'

    def to_pattern_fill(self) -> PatternFill:
        """
        转换为openpyxl的PatternFill对象

        将当前的填充颜色配置转换为openpyxl库使用的PatternFill对象。
        PatternFill用于设置Excel单元格的填充样式。

        返回:
            PatternFill: openpyxl的填充样式对象，配置为纯色填充（solid）

        说明:
            使用start_color和end_color设置为相同的颜色值，创建纯色填充效果。
            fill_type设置为'solid'表示纯色填充。
        """
        return PatternFill(
            start_color=self.fill_color,
            end_color=self.fill_color,
            fill_type='solid'
        )

    def to_font(self) -> Font:
        """
        转换为openpyxl的Font对象

        将当前的字体颜色配置转换为openpyxl库使用的Font对象。
        Font用于设置Excel单元格中文字的字体样式。

        返回:
            Font: openpyxl的字体对象，配置了指定的字体颜色

        说明:
            当前版本只设置字体颜色，其他字体属性（如字体名、大小、粗体等）保持默认。
        """
        return Font(color=self.font_color)


class CellStyles:
    """
    预定义的单元格样式集合

    提供了一组预定义的单元格样式，用于不同类型的匹配结果标记。
    这些样式在系统中用于直观地展示数据的匹配状态。

    样式说明:
        YELLOW (黄色): 标记重复数据。优先级最高，表示该数据在之前已经处理过。
        PURPLE (紫色): 标记日期范围数据且全部匹配成功。表示所有月份的数据都找到了对应记录。
        BROWN (棕色): 标记日期范围数据但未全部匹配成功。表示部分月份的数据未找到对应记录。
        GREEN (绿色): 标记单条数据匹配成功。表示在匹配原表中找到了对应记录。
        RED (红色): 标记单条数据未匹配成功。表示在匹配原表中未找到对应记录。

    使用示例:
        >>> # 获取重复数据样式
        >>> duplicate_style = CellStyles.YELLOW
        >>> fill = duplicate_style.to_pattern_fill()

        >>> # 获取匹配成功样式
        >>> matched_style = CellStyles.GREEN
        >>> fill = matched_style.to_pattern_fill()

    颜色方案:
        - 黄色 ('FFFF00'): 重复数据警告
        - 紫色 ('9370DB') + 白色字体: 日期范围全部匹配成功
        - 棕色 ('8B4513') + 白色字体: 日期范围部分匹配成功
        - 绿色 ('90EE90'): 单条数据匹配成功
        - 红色 ('FFB6C1'): 单条数据匹配失败
    """

    YELLOW = CellStyle('FFFF00')
    PURPLE = CellStyle('9370DB', 'FFFFFF')
    BROWN = CellStyle('8B4513', 'FFFFFF')
    GREEN = CellStyle('90EE90')
    RED = CellStyle('FFB6C1')
