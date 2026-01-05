"""
供应商数据智能匹配系统 - 增强版界面
优化版本 - 简洁现代风格 + 改进交互体验
"""

import sys
import os
import re
import logging
from datetime import datetime
from typing import Tuple, List, Dict, Set, Optional
from dataclasses import dataclass

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QPushButton,
    QVBoxLayout, QHBoxLayout, QTabWidget, QLabel,
    QLineEdit, QFileDialog, QProgressDialog, QMessageBox,
    QProgressBar, QCheckBox, QScrollArea, QFrame,
    QGroupBox, QSizePolicy
)
from PySide6.QtCore import Qt, QSettings, QMimeData, QSize, Signal
from PySide6.QtGui import QIcon, QDragEnterEvent, QDropEvent, QPalette, QColor
import openpyxl
from openpyxl.styles import PatternFill, Font


# ==================== 数据类和常量 ====================

@dataclass
class MatchResult:
    """匹配结果数据类"""
    is_duplicate: bool = False
    is_date_range: bool = False
    is_all_match: bool = False
    is_match: bool = False
    matched_suppliers: List[Tuple[str, str]] = None

    def __post_init__(self):
        if self.matched_suppliers is None:
            self.matched_suppliers = []


@dataclass
class CellStyle:
    """单元格样式配置"""
    fill_color: str
    font_color: str = '000000'

    def to_pattern_fill(self) -> PatternFill:
        return PatternFill(
            start_color=self.fill_color,
            end_color=self.fill_color,
            fill_type='solid'
        )

    def to_font(self) -> Font:
        return Font(color=self.font_color)


class CellStyles:
    """预定义的单元格样式"""
    YELLOW = CellStyle('FFFF00')
    PURPLE = CellStyle('9370DB', 'FFFFFF')
    BROWN = CellStyle('8B4513', 'FFFFFF')
    GREEN = CellStyle('90EE90')
    RED = CellStyle('FFB6C1')


# ==================== 辅助函数 ====================

def standardize_data(value: str, column_index: int) -> str:
    """标准化数据"""
    if not value:
        return ""
    value = ''.join(value.split())

    if column_index == 1:
        return _standardize_date(value)
    elif column_index == 2:
        return _standardize_customer_name(value)
    elif column_index == 3:
        return _standardize_product_name(value)
    return value


def _standardize_date(value: str) -> str:
    """标准化日期数据"""
    logging.debug(f"处理日期值: {value}")
    cn_num_map = {'一': '1', '二': '2', '三': '3', '四': '4', '五': '5',
                  '六': '6', '七': '7', '八': '8', '九': '9', '十': '10', '正': '1'}
    for cn, num in cn_num_map.items():
        value = value.replace(cn, num)

    date_range = _parse_date_range(value)
    if date_range:
        return date_range

    value = value.replace('月', '').replace('年', '')
    date_patterns = [
        (r'(\d{4})[-/.]?(\d{1,2})', 2),
        (r'(\d{2})(\d{2})', 2),
        (r'(\d{1,2})', 1),
    ]

    for pattern, group_count in date_patterns:
        match = re.match(pattern, value)
        if match:
            try:
                groups = match.groups()
                if group_count == 2:
                    year, month = groups
                    if len(year) == 2:
                        year = '20' + year
                else:
                    year = str(datetime.now().year)
                    month = groups[0]

                month = int(month)
                if 1 <= month <= 12:
                    month = str(month).zfill(2)
                    result = f"{year}{month}"
                    logging.debug(f"日期标准化结果: {result}")
                    return result
            except (ValueError, IndexError):
                pass

    logging.debug(f"日期标准化结果: {value} (未改变)")
    return value


def _parse_date_range(value: str) -> Optional[str]:
    """解析日期范围"""
    cn_range_patterns = [
        r'(\d{2,4})年(\d{1,2})月[到至和-](\d{1,2})月',
        r'(\d{2,4})年(\d{1,2})[到至和-](\d{1,2})月',
    ]

    for pattern in cn_range_patterns:
        match = re.search(pattern, value)
        if match:
            year = match.group(1)
            if len(year) == 2:
                year = '20' + year
            start_month = int(match.group(2))
            end_month = int(match.group(3))
            if 1 <= start_month <= 12 and 1 <= end_month <= 12:
                months = [f"{year}{str(m).zfill(2)}" for m in range(start_month, end_month + 1)]
                return ",".join(months)

    num_range_pattern = r'(\d{4})(\d{1,2})-(\d{1,2})'
    match = re.search(num_range_pattern, value)
    if match:
        year = match.group(1)
        start_month = int(match.group(2))
        end_month = int(match.group(3))
        if 1 <= start_month <= 12 and 1 <= end_month <= 12:
            months = [f"{year}{str(m).zfill(2)}" for m in range(start_month, end_month + 1)]
            return ",".join(months)

    return None


def _standardize_customer_name(value: str) -> str:
    """标准化客户名称"""
    value = value.replace('（', '(').replace('）', ')')
    value = value.replace('：', ':').replace('，', ',')
    value = value.replace('"', '"').replace('"', '"')
    value = value.replace('　', '')
    return value


def _standardize_product_name(value: str) -> str:
    """标准化产品名称"""
    value = value.replace('（', '(').replace('）', ')')
    value = value.replace('，', ',').replace('：', ':')
    value = value.replace('　', '')
    return value.upper()


def get_sheet_data(sheet, row: int) -> Tuple[str, str, str]:
    """获取并标准化工作表数据"""
    values = tuple(
        standardize_data(str(sheet.cell(row=row, column=i).value), i)
        for i in range(1, 4)
    )
    logging.debug(f"行{row}原始数据: {values}")
    return values


def clear_sheet(sheet) -> None:
    """清空工作表数据"""
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row)


def copy_title_row(source_sheet, target_sheet) -> None:
    """复制标题行"""
    for column, cell in enumerate(source_sheet[1], start=1):
        target_sheet.cell(row=1, column=column, value=cell.value)


def init_result_sheet(workbook, sheet_name: str):
    """初始化结果工作表"""
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        clear_sheet(sheet)
    else:
        sheet = workbook.create_sheet(sheet_name)
    return sheet


def setup_logging(log_dir: str) -> str:
    """设置日志系统"""
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)

    today = datetime.now().strftime("%Y%m%d")
    try:
        for log_file_name in os.listdir(log_dir):
            if log_file_name.startswith('供应商匹配_') and log_file_name.endswith('.log'):
                file_date = log_file_name.replace('供应商匹配_', '').replace('.log', '')
                if file_date <= today:
                    old_log_path = os.path.join(log_dir, log_file_name)
                    try:
                        os.remove(old_log_path)
                        print(f"已清理历史日志: {log_file_name}")
                    except Exception as e:
                        print(f"清理日志文件失败 {log_file_name}: {str(e)}")
    except Exception as e:
        print(f"清理历史日志时出错: {str(e)}")

    log_file = os.path.join(log_dir, f'供应商匹配_{today}.log')
    if os.path.exists(log_file):
        try:
            os.remove(log_file)
        except Exception as e:
            print(f"清理旧日志文件失败: {str(e)}")

    logging.basicConfig(
        filename=log_file,
        level=logging.DEBUG,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )

    return log_file


# ==================== 自定义组件 ====================

class DropZoneWidget(QFrame):
    """支持拖拽的文件选择区域"""

    # 定义信号
    file_selected = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self._setup_ui()

    def _setup_ui(self):
        """设置UI"""
        self.setFrameStyle(QFrame.Box)
        self.setStyleSheet("""
            QFrame {
                border: 2px dashed #CCCCCC;
                border-radius: 8px;
                background-color: #FAFAFA;
                padding: 20px;
            }
            QFrame:hover {
                border-color: #4CAF50;
                background-color: #F0F8F0;
            }
        """)

        layout = QVBoxLayout()
        layout.setAlignment(Qt.AlignCenter)

        # 图标和提示文字
        icon_label = QLabel("📁")
        icon_label.setStyleSheet("font-size: 48px;")
        icon_label.setAlignment(Qt.AlignCenter)

        title_label = QLabel("拖拽Excel文件到这里")
        title_label.setObjectName("title_label")
        title_label.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: bold;
                color: #546E7A;
                padding: 10px 0px;
            }
        """)
        title_label.setAlignment(Qt.AlignCenter)

        subtitle_label = QLabel("或者点击下方按钮选择文件")
        subtitle_label.setObjectName("subtitle_label")
        subtitle_label.setStyleSheet("""
            QLabel {
                font-size: 12px;
                color: #607D8B;
                padding: 5px 0px;
            }
        """)
        subtitle_label.setAlignment(Qt.AlignCenter)

        layout.addWidget(icon_label)
        layout.addWidget(title_label)
        layout.addWidget(subtitle_label)
        self.setLayout(layout)

    def dragEnterEvent(self, event: QDragEnterEvent):
        """拖拽进入事件"""
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            self.setStyleSheet("""
                QFrame {
                    border: 2px dashed #4CAF50;
                    border-radius: 8px;
                    background-color: #E8F5E9;
                    padding: 20px;
                }
            """)

    def dragLeaveEvent(self, event):
        """拖拽离开事件"""
        self.setStyleSheet("""
            QFrame {
                border: 2px dashed #CCCCCC;
                border-radius: 8px;
                background-color: #FAFAFA;
                padding: 20px;
            }
        """)

    def dropEvent(self, event: QDropEvent):
        """拖拽放下事件"""
        files = [u.toLocalFile() for u in event.mimeData().urls()]
        if files:
            file_path = files[0]
            if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
                self.file_selected.emit(file_path)
                self._update_display(file_path)
            else:
                QMessageBox.warning(self, "警告", "请选择Excel文件(.xlsx或.xls)")
        self._reset_style()

    def _reset_style(self):
        """重置样式"""
        self.setStyleSheet("""
            QFrame {
                border: 2px dashed #CCCCCC;
                border-radius: 8px;
                background-color: #FAFAFA;
                padding: 20px;
            }
        """)

    def _update_display(self, file_path: str):
        """更新显示文件信息"""
        file_name = os.path.basename(file_path)
        file_size = os.path.getsize(file_path)
        file_size_mb = file_size / (1024 * 1024)  # 转换为MB

        # 更新标题 - 显示文件名
        title = self.findChild(QLabel, "title_label")
        if title:
            title.setText(f"✅ {file_name}")
            title.setStyleSheet("""
                QLabel {
                    font-size: 16px;
                    font-weight: bold;
                    color: #4CAF50;
                    padding: 10px 0px;
                }
            """)

        # 更新副标题 - 显示文件路径和大小
        subtitle = self.findChild(QLabel, "subtitle_label")
        if subtitle:
            subtitle.setText(f"📂 {file_path}\n📊 文件大小: {file_size_mb:.2f} MB")
            subtitle.setStyleSheet("""
                QLabel {
                    font-size: 11px;
                    color: #4CAF50;
                    padding: 5px 0px;
                }
            """)


class DropZoneGroupBox(QGroupBox):
    """支持拖拽的文件选择GroupBox - 整个卡片都支持拖放"""

    # 定义信号
    file_selected = Signal(str)

    def __init__(self, title: str, parent=None):
        super().__init__(title, parent)
        self.setAcceptDrops(True)
        self.current_file_path = ""
        self._setup_ui()

    def _setup_ui(self):
        """设置UI"""
        layout = QVBoxLayout(self)
        layout.setSpacing(15)

        # 拖拽区域内容 - 添加到主布局
        drag_content = QWidget()
        drag_layout = QVBoxLayout(drag_content)
        drag_layout.setAlignment(Qt.AlignCenter)
        drag_layout.setSpacing(10)

        # 图标和提示文字
        icon_label = QLabel("📁")
        icon_label.setStyleSheet("font-size: 64px;")
        icon_label.setAlignment(Qt.AlignCenter)

        title_label = QLabel("拖拽Excel文件到这里")
        title_label.setObjectName("drop_title_label")
        title_label.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                color: #546E7A;
                padding: 10px 0px;
            }
        """)
        title_label.setAlignment(Qt.AlignCenter)

        subtitle_label = QLabel("整个卡片都支持拖放文件")
        subtitle_label.setObjectName("drop_subtitle_label")
        subtitle_label.setStyleSheet("""
            QLabel {
                font-size: 13px;
                color: #607D8B;
                padding: 5px 0px;
            }
        """)
        subtitle_label.setAlignment(Qt.AlignCenter)

        drag_layout.addWidget(icon_label)
        drag_layout.addWidget(title_label)
        drag_layout.addWidget(subtitle_label)

        # 设置整个GroupBox的样式
        self.setStyleSheet("""
            QGroupBox {
                border: 2px dashed #CCCCCC;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 20px;
                font-weight: bold;
                background-color: #FAFAFA;
            }
            QGroupBox:hover {
                border-color: #4CAF50;
                background-color: #F0F8F0;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: #607D8B;
            }
        """)

        self.drag_content = drag_content
        layout.addWidget(drag_content)

        # 按钮布局容器
        self.button_container = QWidget()
        layout.addWidget(self.button_container)

    def add_button_layout(self, button_layout):
        """添加按钮布局"""
        container_layout = QVBoxLayout(self.button_container)
        container_layout.addLayout(button_layout)

    def dragEnterEvent(self, event: QDragEnterEvent):
        """拖拽进入事件"""
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            self.setStyleSheet("""
                QGroupBox {
                    border: 3px dashed #4CAF50;
                    border-radius: 8px;
                    margin-top: 10px;
                    padding-top: 20px;
                    font-weight: bold;
                    background-color: #E8F5E9;
                }
                QGroupBox::title {
                    subcontrol-origin: margin;
                    left: 10px;
                    padding: 0 5px;
                    color: #4CAF50;
                }
            """)

    def dragLeaveEvent(self, event):
        """拖拽离开事件"""
        self.setStyleSheet("""
            QGroupBox {
                border: 2px dashed #CCCCCC;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 20px;
                font-weight: bold;
                background-color: #FAFAFA;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: #607D8B;
            }
        """)

    def dropEvent(self, event: QDropEvent):
        """拖拽放下事件"""
        files = [u.toLocalFile() for u in event.mimeData().urls()]
        if files:
            file_path = files[0]
            if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
                self.file_selected.emit(file_path)
                self._update_display(file_path)
            else:
                QMessageBox.warning(self, "警告", "请选择Excel文件(.xlsx或.xls)")
        self._reset_style()

    def _reset_style(self):
        """重置样式"""
        self.setStyleSheet("""
            QGroupBox {
                border: 2px dashed #CCCCCC;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 20px;
                font-weight: bold;
                background-color: #FAFAFA;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: #607D8B;
            }
        """)

    def _update_display(self, file_path: str):
        """更新显示文件信息"""
        self.current_file_path = file_path
        file_name = os.path.basename(file_path)
        file_size = os.path.getsize(file_path)
        file_size_mb = file_size / (1024 * 1024)  # 转换为MB

        # 更新标题 - 显示文件名
        title = self.findChild(QLabel, "drop_title_label")
        if title:
            title.setText(f"✅ {file_name}")
            title.setStyleSheet("""
                QLabel {
                    font-size: 18px;
                    font-weight: bold;
                    color: #4CAF50;
                    padding: 10px 0px;
                }
            """)

        # 更新副标题 - 显示文件路径和大小
        subtitle = self.findChild(QLabel, "drop_subtitle_label")
        if subtitle:
            subtitle.setText(f"📂 {file_path}\n📊 文件大小: {file_size_mb:.2f} MB")
            subtitle.setStyleSheet("""
                QLabel {
                    font-size: 12px;
                    color: #4CAF50;
                    padding: 5px 0px;
                }
            """)


class StatCard(QFrame):
    """统计卡片组件"""

    def __init__(self, title: str, value: str = "0", icon: str = "📊", description: str = "", parent=None):
        super().__init__(parent)
        self.title = title
        self.value = value
        self.icon = icon
        self.description = description
        self._setup_ui()

    def _setup_ui(self):
        """设置UI"""
        self.setFrameStyle(QFrame.Box)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.setStyleSheet("""
            QFrame {
                border: 1px solid #E0E0E0;
                border-radius: 8px;
                background-color: white;
                padding: 15px;
            }
            QFrame:hover {
                background-color: #FAFAFA;
            }
        """)

        layout = QVBoxLayout()
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(10)

        # 图标和数值
        header_layout = QHBoxLayout()

        icon_label = QLabel(self.icon)
        icon_label.setStyleSheet("font-size: 24px;")

        value_label = QLabel(self.value)
        value_label.setObjectName("value_label")
        value_label.setStyleSheet("""
            QLabel#value_label {
                font-size: 28px;
                font-weight: bold;
                color: #4CAF50;
            }
        """)
        value_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)

        header_layout.addWidget(icon_label)
        header_layout.addStretch()
        header_layout.addWidget(value_label)

        layout.addLayout(header_layout)

        # 标题
        title_label = QLabel(self.title)
        title_label.setStyleSheet("""
            QLabel {
                font-size: 13px;
                font-weight: bold;
                color: #546E7A;
            }
        """)
        layout.addWidget(title_label)

        # 描述说明 - 使用更明显的样式
        if self.description:
            desc_label = QLabel(self.description)
            desc_label.setWordWrap(True)
            desc_label.setAlignment(Qt.AlignLeft)
            desc_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            desc_label.setStyleSheet("""
                QLabel {
                    font-size: 16px;
                    font-weight: bold;
                    color: #0D47A1;
                    background-color: #E3F2FD;
                    padding: 15px;
                    border-radius: 8px;
                    border: 2px solid #BBDEFB;
                }
            """)
            layout.addWidget(desc_label, 1)  # stretch=1 让描述占据剩余空间

            # 同时添加工具提示
            self.setToolTip(self.description)

        self.setLayout(layout)

    def update_value(self, value: str):
        """更新数值"""
        value_label = self.findChild(QLabel, "value_label")
        if value_label:
            value_label.setText(value)


# ==================== 主窗口类 ====================

class MainWindow(QMainWindow):
    """供应商数据智能匹配系统主窗口 - 增强版"""

    def __init__(self):
        super().__init__()
        self.settings = QSettings('供应商数据智能匹配系统', 'DataAnalysis')
        self.log_file: Optional[str] = None
        self.recent_files: List[str] = []

        self._init_logging()
        self._init_ui()

        logging.info("系统初始化完成")

    def _init_logging(self):
        """初始化日志系统"""
        log_dir = os.path.join(os.path.dirname(__file__), 'logs')
        self.log_file = setup_logging(log_dir)

        settings = QSettings('供应商数据智能匹配系统', 'DataAnalysis')
        enable_logging = settings.value('enable_logging', False, bool)
        logging.getLogger().disabled = not enable_logging

    def _init_ui(self):
        """初始化用户界面"""
        self.setWindowTitle("供应商数据智能匹配系统")
        self._set_window_icon()
        self.setMinimumSize(800, 600)

        # 设置应用样式
        self._set_app_style()

        # 创建中央部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # 创建标签页
        tab_widget = QTabWidget()
        tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: none;
                background-color: transparent;
            }
            QTabBar::tab {
                background-color: #F5F5F5;
                color: #607D8B;
                padding: 10px 20px;
                margin-right: 5px;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
                font-size: 13px;
            }
            QTabBar::tab:selected {
                background-color: white;
                color: #4CAF50;
                font-weight: bold;
            }
            QTabBar::tab:hover:!selected {
                background-color: #EEEEEE;
                color: #546E7A;
            }
        """)
        main_layout.addWidget(tab_widget)

        # 添加标签页
        tab_widget.addTab(self._create_filter_tab(), "📊 数据筛选")
        tab_widget.addTab(self._create_settings_tab(), "⚙️ 设置")

        # 添加底部按钮
        self._create_bottom_buttons(main_layout)

    def _set_app_style(self):
        """设置应用程序样式"""
        QApplication.setStyle("Fusion")

        # 全局样式 - 优化配色，减少黑色，添加苹果风格滚动条
        self.setStyleSheet("""
            QMainWindow {
                background-color: #F5F5F5;
            }
            QWidget {
                font-family: "Microsoft YaHei", "PingFang SC", sans-serif;
                font-size: 13px;
                color: #546E7A;
            }
            /* 苹果风格滚动条 - 垂直 */
            QScrollBar:vertical {
                border: none;
                background: transparent;
                width: 10px;
                margin: 0px;
            }
            QScrollBar::handle:vertical {
                background: #C1C1C1;
                min-height: 30px;
                border-radius: 5px;
                margin: 2px;
            }
            QScrollBar::handle:vertical:hover {
                background: #A8A8A8;
            }
            QScrollBar::handle:vertical:pressed {
                background: #8F8F8F;
            }
            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {
                border: none;
                background: none;
                height: 0px;
            }
            QScrollBar::add-page:vertical,
            QScrollBar::sub-page:vertical {
                background: none;
            }
            /* 苹果风格滚动条 - 水平 */
            QScrollBar:horizontal {
                border: none;
                background: transparent;
                height: 10px;
                margin: 0px;
            }
            QScrollBar::handle:horizontal {
                background: #C1C1C1;
                min-width: 30px;
                border-radius: 5px;
                margin: 2px;
            }
            QScrollBar::handle:horizontal:hover {
                background: #A8A8A8;
            }
            QScrollBar::handle:horizontal:pressed {
                background: #8F8F8F;
            }
            QScrollBar::add-line:horizontal,
            QScrollBar::sub-line:horizontal {
                border: none;
                background: none;
                width: 0px;
            }
            QScrollBar::add-page:horizontal,
            QScrollBar::sub-page:horizontal {
                background: none;
            }
            QScrollArea {
                border: none;
                background-color: transparent;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 6px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #3d8b40;
            }
            QPushButton:disabled {
                background-color: #E0E0E0;
                color: #9E9E9E;
            }
            QLineEdit {
                padding: 8px 12px;
                border: 1px solid #CFD8DC;
                border-radius: 4px;
                background-color: white;
                font-size: 13px;
                color: #546E7A;
            }
            QLineEdit:focus {
                border: 1px solid #4CAF50;
            }
            QProgressBar {
                border: none;
                border-radius: 4px;
                background-color: #E0E0E0;
                text-align: center;
                font-size: 12px;
                color: #546E7A;
            }
            QProgressBar::chunk {
                background-color: #4CAF50;
                border-radius: 4px;
            }
            QGroupBox {
                border: 1px solid #E0E0E0;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
                font-weight: bold;
                background-color: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: #607D8B;
            }
        """)

    def _set_window_icon(self):
        """设置窗口图标"""
        icon_path = os.path.join(os.path.dirname(__file__), 'resources', 'icon.ico')
        if getattr(sys, 'frozen', False):
            base_path = sys._MEIPASS
            icon_path = os.path.join(base_path, 'resources', 'icon.ico')

        if os.path.exists(icon_path):
            app_icon = QIcon(icon_path)
            self.setWindowIcon(app_icon)
            QApplication.setWindowIcon(app_icon)

    def _create_filter_tab(self) -> QWidget:
        """创建数据筛选标签页"""
        # 创建滚动区域
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll_area.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        # 创建容器 - 设置与整体背景融合的颜色
        container = QWidget()
        container.setStyleSheet("background-color: #FAFAFA;")  # 浅灰色背景，与整体更融合
        layout = QVBoxLayout(container)
        layout.setSpacing(20)

        # 添加帮助区域（简化版）
        layout.addWidget(self._create_compact_help())

        # 文件选择区域 - 整个卡片支持拖拽
        self.file_group = DropZoneGroupBox("📁 文件选择")
        self.file_group.setMinimumHeight(300)  # 设置更大的最小高度
        self.file_group.file_selected.connect(self._on_file_dropped)

        # 按钮区域 - 浏览文件和开始分析并排
        button_layout = QHBoxLayout()
        button_layout.addStretch()

        browse_button = QPushButton("📂 浏览文件")
        browse_button.clicked.connect(self.browse_file)
        browse_button.setMinimumHeight(45)
        browse_button.setMinimumWidth(150)
        browse_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 12px 24px;
                border-radius: 8px;
                font-weight: bold;
                font-size: 15px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #3d8b40;
            }
        """)
        button_layout.addWidget(browse_button)

        # 开始分析按钮
        self.analyze_button = QPushButton("🚀 开始分析")
        self.analyze_button.clicked.connect(self.start_analysis)
        self.analyze_button.setEnabled(False)  # 初始状态不可点击
        self.analyze_button.setMinimumHeight(45)
        self.analyze_button.setMinimumWidth(150)
        self.analyze_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 12px 24px;
                border-radius: 8px;
                font-weight: bold;
                font-size: 15px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:disabled {
                background-color: #E0E0E0;
                color: #9E9E9E;
                border: 1px solid #D0D0D0;
            }
        """)
        button_layout.addWidget(self.analyze_button)
        button_layout.addStretch()

        self.file_group.add_button_layout(button_layout)
        layout.addWidget(self.file_group)

        # 统计信息卡片 - 响应式布局，不使用滚动条
        stats_group = QGroupBox("📊 分析统计")
        # 设置统计组的大小策略和最小高度
        stats_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        stats_group.setMinimumHeight(500)  # 增加最小高度，确保内容全部展示

        # 创建统计卡片布局 - 直接使用VBoxLayout，不用滚动区域
        stats_layout = QVBoxLayout(stats_group)
        stats_layout.setSpacing(15)
        stats_layout.setContentsMargins(10, 20, 10, 10)

        # 创建第一行卡片(水平布局)
        first_row = QWidget()
        first_row_layout = QHBoxLayout(first_row)
        first_row_layout.setSpacing(15)

        # 创建第二行卡片(水平布局)
        second_row = QWidget()
        second_row_layout = QHBoxLayout(second_row)
        second_row_layout.setSpacing(15)

        # 创建统计卡片
        self.stat_total = StatCard("总数据", "0", "📋", "待处理的数据总数")
        self.stat_matched = StatCard("已匹配", "0", "✅", "成功匹配到的数据条数")
        self.stat_unmatched = StatCard("未匹配", "0", "❌", "未找到对应的数据条数")
        self.stat_rate = StatCard("匹配率", "0%", "📈", "成功匹配的百分比")

        # 为每个卡片设置最小高度,确保描述文字有足够空间
        self.stat_total.setMinimumHeight(220)
        self.stat_matched.setMinimumHeight(220)
        self.stat_unmatched.setMinimumHeight(220)
        self.stat_rate.setMinimumHeight(220)

        # 设置伸展因子
        self.stat_total.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.stat_matched.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.stat_unmatched.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.stat_rate.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        # 添加到行布局
        first_row_layout.addWidget(self.stat_total)
        first_row_layout.addWidget(self.stat_matched)
        second_row_layout.addWidget(self.stat_unmatched)
        second_row_layout.addWidget(self.stat_rate)

        # 添加行到主布局
        stats_layout.addWidget(first_row)
        stats_layout.addWidget(second_row)
        stats_layout.addStretch()  # 添加弹性空间

        # 添加统计组到主布局,并设置伸展因子
        layout.addWidget(stats_group, 3)  # stretch=3 让统计区域占据更多垂直空间

        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setFixedHeight(25)
        layout.addWidget(self.progress_bar)

        # 当前选择的文件路径（隐藏）
        self.current_file_path = ""

        # 将容器放入滚动区域
        scroll_area.setWidget(container)

        # 创建主标签页并返回
        tab = QWidget()
        tab.setStyleSheet("background-color: #FAFAFA;")  # 与容器背景一致
        main_layout = QVBoxLayout(tab)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll_area)

        return tab

    def _create_compact_help(self) -> QWidget:
        """创建精简版帮助提示"""
        widget = QFrame()
        widget.setStyleSheet("""
            QFrame {
                background-color: #E3F2FD;
                border: 1px solid #BBDEFB;
                border-radius: 6px;
                padding: 10px;
            }
        """)

        layout = QHBoxLayout(widget)
        layout.setContentsMargins(15, 10, 15, 10)

        icon_label = QLabel("💡")
        icon_label.setStyleSheet("font-size: 20px;")

        text_label = QLabel(
            "将包含两个工作表的Excel文件拖放到上方区域，第一个为待匹配表，第二个为匹配原表"
        )
        text_label.setWordWrap(True)
        text_label.setStyleSheet("color: #1976D2; font-size: 12px;")

        toggle_button = QPushButton("查看详情")
        toggle_button.setCheckable(True)
        toggle_button.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: #1976D2;
                border: 1px solid #1976D2;
                padding: 5px 15px;
                border-radius: 4px;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #BBDEFB;
            }
        """)
        toggle_button.clicked.connect(self._show_detailed_help)

        layout.addWidget(icon_label)
        layout.addWidget(text_label, 1)
        layout.addWidget(toggle_button)

        return widget

    def _show_detailed_help(self):
        """显示详细帮助"""
        help_text = self._get_help_text()
        QMessageBox.information(
            self,
            "使用说明",
            help_text,
            QMessageBox.Ok
        )

    def _get_help_text(self) -> str:
        """获取帮助文本内容"""
        return """使用说明：

1. 📁 数据准备：
   • 第一个工作表为"供应商待匹配表"，放入需要查询的数据
   • 第二个工作表为"供应商匹配原表"，放入用于匹配的数据
   • 两个工作表的前三列必须包含：日期、客户名称、产品名称

2. 📅 数据格式要求：
   • 日期格式支持：2024-03、24年3月、3月、202411-12
   • 客户名称：不区分全角半角，自动处理空格
   • 产品名称：不区分大小写，自动处理特殊符号

3. 🎨 处理结果说明：
   • 🟩绿色：在匹配原表中找到对应数据
   • 🟥红色：在匹配原表中未找到对应数据
   • 🟨黄色：该数据重复查询（最高优先级）
   • 🟫棕色：日期范围内的数据未能全部匹配成功
   • 🟪紫色：日期范围内的数据全部匹配成功

4. 💡 使用技巧：
   • 可以直接拖拽Excel文件到窗口
   • 支持批量处理大量数据
   • 分析结果会自动保存到原文件"""

    def _on_file_dropped(self, file_path: str):
        """处理拖拽的文件"""
        if os.path.exists(file_path):
            self.current_file_path = file_path
            self.analyze_button.setEnabled(True)
            logging.info(f"通过拖拽选择文件: {file_path}")
        else:
            QMessageBox.warning(self, "警告", "文件不存在")

    def _create_settings_tab(self) -> QWidget:
        """创建设置标签页"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(20)
        layout.setContentsMargins(20, 20, 20, 20)

        # 日志设置
        log_group = QGroupBox("📝 日志设置")
        log_layout = QVBoxLayout(log_group)

        log_checkbox = QCheckBox("启用日志记录")
        log_checkbox.setChecked(self.settings.value('enable_logging', False, bool))
        log_checkbox.stateChanged.connect(self.toggle_logging)
        log_layout.addWidget(log_checkbox)

        log_path_label = QLabel(f"日志文件位置：{os.path.abspath(self.log_file)}")
        log_path_label.setWordWrap(True)
        log_path_label.setStyleSheet("color: #666666; font-size: 11px;")
        log_layout.addWidget(log_path_label)

        layout.addWidget(log_group)

        # 关于信息
        about_group = QGroupBox("ℹ️ 关于")
        about_layout = QVBoxLayout(about_group)

        about_text = QLabel(
            "供应商数据智能匹配系统 v1.0\n\n"
            "一个现代化的Excel数据匹配工具，\n"
            "帮助您快速处理和分析供应商数据。\n\n"
            "特性：\n"
            "• 支持拖拽上传\n"
            "• 智能日期范围处理\n"
            "• 实时统计显示\n"
            "• 简洁现代界面"
        )
        about_text.setWordWrap(True)
        about_layout.addWidget(about_text)

        layout.addWidget(about_group)
        layout.addStretch()

        return tab

    def _create_bottom_buttons(self, layout: QVBoxLayout):
        """创建底部按钮"""
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)

        button_layout.addStretch()

        exit_button = QPushButton("退出")
        exit_button.setStyleSheet("""
            QPushButton {
                background-color: #F5F5F5;
                color: #666666;
                border: 1px solid #DDDDDD;
                padding: 8px 20px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #EEEEEE;
            }
        """)
        exit_button.clicked.connect(self.close)
        button_layout.addWidget(exit_button)

        layout.addLayout(button_layout)

    def browse_file(self):
        """浏览并选择Excel文件"""
        logging.info("开始选择文件")
        try:
            last_dir = self.settings.value('last_directory', os.path.expanduser("~/Documents"))

            file_name, _ = QFileDialog.getOpenFileName(
                self,
                "选择Excel文件",
                last_dir,
                "Excel Files (*.xlsx *.xls);;All Files (*.*)"
            )

            if file_name:
                self._validate_and_set_file(file_name)
            else:
                logging.info("未选择文件")

        except Exception as e:
            logging.error(f"选择文件时出错: {str(e)}")
            QMessageBox.critical(self, "错误", f"选择文件时出错：{str(e)}")

    def _validate_and_set_file(self, file_name: str):
        """验证并设置文件路径"""
        if not os.path.exists(file_name):
            logging.error(f"文件不存在: {file_name}")
            QMessageBox.critical(self, "错误", "所选文件不存在")
            return

        if not os.access(file_name, os.R_OK):
            logging.error(f"文件无法访问: {file_name}")
            QMessageBox.critical(self, "错误", "无法访问所选文件")
            return

        self.settings.setValue('last_directory', os.path.dirname(file_name))
        self.current_file_path = file_name
        self.analyze_button.setEnabled(True)
        logging.info(f"选择的文件: {file_name}")

        # 更新文件选择区域的显示
        self.file_group._update_display(file_name)

    def toggle_logging(self, state: int):
        """切换日志记录状态"""
        self.settings.setValue('enable_logging', bool(state))
        logging.getLogger().disabled = not state
        status = "启用" if state else "禁用"
        logging.info(f"日志记录已{status}")

    def start_analysis(self):
        """开始数据分析"""
        if not self.current_file_path:
            QMessageBox.warning(self, "警告", "请先选择Excel文件")
            return

        logging.info("开始数据分析")
        self.analyze_button.setEnabled(False)

        try:
            selected_file = self.current_file_path
            workbook = openpyxl.load_workbook(selected_file)
            logging.info(f"工作簿包含的工作表: {workbook.sheetnames}")

            if len(workbook.worksheets) < 2:
                QMessageBox.critical(self, "错误", "工作簿中缺少必要的工作表")
                self.analyze_button.setEnabled(True)
                return

            # 获取工作表
            sheet1 = workbook.worksheets[0]
            sheet2 = workbook.worksheets[1]
            sheet3 = init_result_sheet(workbook, "匹配到的数据")
            sheet4 = init_result_sheet(workbook, "未找到的数据")

            # 显示进度条
            self.progress_bar.setVisible(True)

            # 处理数据
            stats = self.process_data(workbook, sheet1, sheet2, sheet3, sheet4)

            # 保存结果
            workbook.save(selected_file)
            self.progress_bar.setVisible(False)

            # 更新统计信息
            self._update_stats(stats)

            logging.info("数据分析完成")
            QMessageBox.information(
                self,
                "✅ 分析完成",
                f"数据处理完成！\n\n"
                f"总计：{stats['total']} 条\n"
                f"已匹配：{stats['matched']} 条\n"
                f"未匹配：{stats['unmatched']} 条\n"
                f"匹配率：{stats['rate']}%",
                QMessageBox.Ok
            )

        except Exception as e:
            self.progress_bar.setVisible(False)
            logging.error(f"分析过程出错: {str(e)}", exc_info=True)
            QMessageBox.critical(self, "错误", f"执行分析时出错：{str(e)}")
        finally:
            self.analyze_button.setEnabled(True)

    def process_data(self, workbook, sheet1, sheet2, sheet3, sheet4):
        """处理数据匹配逻辑"""
        try:
            logging.info("开始处理数据")

            # 初始化结果表
            copy_title_row(sheet1, sheet3)
            copy_title_row(sheet1, sheet4)
            sheet3.cell(row=1, column=4, value="供应商")
            sheet4.cell(row=1, column=4, value="供应商")

            # 预处理匹配数据
            sheet2_data = {}
            for row in sheet2.iter_rows(min_row=2, values_only=True):
                key = (
                    standardize_data(str(row[0]), 1),
                    standardize_data(str(row[1]), 2),
                    standardize_data(str(row[2]), 3)
                )
                if key in sheet2_data:
                    sheet2_data[key].append(row[3])
                else:
                    sheet2_data[key] = [row[3]]

            # 处理数据
            max_row = sheet1.max_row
            if max_row <= 1:
                raise ValueError("Sheet1中没有数据需要匹配")

            self.progress_bar.setMaximum(max_row - 1)

            processed_keys: Set[Tuple[str, str, str]] = set()
            date_range_map: Dict[Tuple[str, str], List[str]] = {}

            matched_count = 0
            unmatched_count = 0

            for row in range(2, max_row + 1):
                self.progress_bar.setValue(row - 2)

                original_data = tuple(str(sheet1.cell(row=row, column=i).value) for i in range(1, 4))
                search_key = get_sheet_data(sheet1, row)

                # 分析匹配
                result = self._analyze_match(search_key, sheet2_data, processed_keys, date_range_map)

                # 应用样式
                cell_style = self._determine_cell_style(result)
                for col in range(1, 4):
                    cell = sheet1.cell(row=row, column=col)
                    cell.fill = cell_style.to_pattern_fill()
                    cell.font = cell_style.to_font()

                # 保存结果
                if result.is_match or (result.is_date_range and result.is_all_match):
                    matched_count += 1
                    target_sheet = sheet3
                    for _, supplier in result.matched_suppliers:
                        target_sheet.append(original_data + (supplier,))
                else:
                    unmatched_count += 1
                    sheet4.append(original_data + ('',))

                processed_keys.add(search_key)

            # 计算统计
            total = matched_count + unmatched_count
            rate = (matched_count / total * 100) if total > 0 else 0

            return {
                'total': total,
                'matched': matched_count,
                'unmatched': unmatched_count,
                'rate': f"{rate:.1f}"
            }

        except Exception as e:
            self.progress_bar.setVisible(False)
            logging.error(f"数据处理出错: {str(e)}", exc_info=True)
            raise

    def _analyze_match(self, search_key: Tuple[str, str, str],
                       sheet2_data: Dict,
                       processed_keys: Set[Tuple[str, str, str]],
                       date_range_map: Dict[Tuple[str, str], List[str]]) -> MatchResult:
        """分析数据匹配情况"""
        result = MatchResult()
        result.is_duplicate = self._check_duplicate(search_key, processed_keys, date_range_map)

        if ',' in search_key[0]:
            result.is_date_range = True
            dates = search_key[0].split(',')
            date_range_map[search_key[1:]] = dates

            all_matches = True
            for date in dates:
                test_key = (date,) + search_key[1:]
                if test_key in sheet2_data:
                    for supplier in sheet2_data[test_key]:
                        result.matched_suppliers.append((date, supplier))
                else:
                    all_matches = False

            result.is_all_match = all_matches and bool(result.matched_suppliers)
        elif not result.is_duplicate and search_key in sheet2_data:
            result.is_match = True
            for supplier in sheet2_data[search_key]:
                result.matched_suppliers.append((search_key[0], supplier))

        return result

    def _check_duplicate(self, search_key: Tuple[str, str, str],
                         processed_keys: Set[Tuple[str, str, str]],
                         date_range_map: Dict[Tuple[str, str], List[str]]) -> bool:
        """检查是否为重复数据"""
        if search_key in processed_keys:
            return True

        if ',' not in search_key[0]:
            for range_key, months in date_range_map.items():
                if search_key[1:] == range_key and search_key[0] in months:
                    return True

        if ',' in search_key[0]:
            dates = search_key[0].split(',')
            for date in dates:
                single_key = (date,) + search_key[1:]
                if single_key in processed_keys:
                    return True

        return False

    def _determine_cell_style(self, result: MatchResult) -> CellStyle:
        """根据匹配结果确定单元格样式"""
        if result.is_duplicate:
            return CellStyles.YELLOW
        elif result.is_date_range:
            return CellStyles.PURPLE if result.is_all_match else CellStyles.BROWN
        elif result.is_match:
            return CellStyles.GREEN
        else:
            return CellStyles.RED

    def _update_stats(self, stats: Dict):
        """更新统计信息"""
        self.stat_total.update_value(str(stats['total']))
        self.stat_matched.update_value(str(stats['matched']))
        self.stat_unmatched.update_value(str(stats['unmatched']))
        self.stat_rate.update_value(f"{stats['rate']}%")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
