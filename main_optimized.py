"""
供应商数据智能匹配系统
优化版本 - 提高代码可读性、可维护性和性能
"""

import sys
import os
import re
import logging
from datetime import datetime
from typing import Tuple, List, Dict, Set, Optional, Any
from dataclasses import dataclass

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QPushButton,
    QVBoxLayout, QHBoxLayout, QTabWidget, QLabel,
    QLineEdit, QFileDialog, QProgressDialog, QMessageBox,
    QProgressBar, QCheckBox, QScrollArea
)
from PySide6.QtCore import Qt, QSettings
from PySide6.QtGui import QIcon
import openpyxl
from openpyxl.styles import PatternFill, Font


# ==================== 数据类和常量 ====================

@dataclass
class MatchResult:
    """匹配结果数据类"""
    is_duplicate: bool = False
    is_date_range: bool = False
    is_all_match: bool = False
    is_match: bool = False
    matched_suppliers: List[Tuple[str, str]] = None  # (date, supplier)

    def __post_init__(self):
        if self.matched_suppliers is None:
            self.matched_suppliers = []


@dataclass
class CellStyle:
    """单元格样式配置"""
    fill_color: str
    font_color: str = '000000'

    def to_pattern_fill(self) -> PatternFill:
        """转换为 PatternFill 对象"""
        return PatternFill(
            start_color=self.fill_color,
            end_color=self.fill_color,
            fill_type='solid'
        )

    def to_font(self) -> Font:
        """转换为 Font 对象"""
        return Font(color=self.font_color)


class CellStyles:
    """预定义的单元格样式"""
    YELLOW = CellStyle('FFFF00')      # 重复数据
    PURPLE = CellStyle('9370DB', 'FFFFFF')  # 日期范围全部匹配
    BROWN = CellStyle('8B4513', 'FFFFFF')   # 日期范围部分匹配
    GREEN = CellStyle('90EE90')       # 单条匹配
    RED = CellStyle('FFB6C1')         # 未匹配


# ==================== 辅助函数 ====================

def standardize_data(value: str, column_index: int) -> str:
    """标准化数据

    Args:
        value: 原始值
        column_index: 列索引 (1=日期, 2=客户名称, 3=产品名称)

    Returns:
        标准化后的值
    """
    if not value:
        return ""

    # 基础清理：去除所有空白字符
    value = ''.join(value.split())

    if column_index == 1:
        return _standardize_date(value)
    elif column_index == 2:
        return _standardize_customer_name(value)
    elif column_index == 3:
        return _standardize_product_name(value)

    return value


def _standardize_date(value: str) -> str:
    """标准化日期数据"""
    logging.debug(f"处理日期值: {value}")

    # 处理中文数字
    cn_num_map = {'一': '1', '二': '2', '三': '3', '四': '4', '五': '5',
                  '六': '6', '七': '7', '八': '8', '九': '9', '十': '10', '正': '1'}
    for cn, num in cn_num_map.items():
        value = value.replace(cn, num)

    # 处理日期范围
    date_range = _parse_date_range(value)
    if date_range:
        return date_range

    # 移除中文字符
    value = value.replace('月', '').replace('年', '')

    # 处理标准日期格式
    date_patterns = [
        (r'(\d{4})[-/.]?(\d{1,2})', 2),  # 2024/4, 2024-04
        (r'(\d{2})(\d{2})', 2),          # 2404
        (r'(\d{1,2})', 1),               # 单独月份
    ]

    for pattern, group_count in date_patterns:
        match = re.match(pattern, value)
        if match:
            try:
                groups = match.groups()
                if group_count == 2:
                    year, month = groups
                    if len(year) == 2:
                        year = '20' + year
                else:
                    year = str(datetime.now().year)
                    month = groups[0]

                month = int(month)
                if 1 <= month <= 12:
                    month = str(month).zfill(2)
                    result = f"{year}{month}"
                    logging.debug(f"日期标准化结果: {result}")
                    return result
            except (ValueError, IndexError):
                pass

    logging.debug(f"日期标准化结果: {value} (未改变)")
    return value


def _parse_date_range(value: str) -> Optional[str]:
    """解析日期范围，返回逗号分隔的月份列表"""
    # 处理中文日期范围
    cn_range_patterns = [
        r'(\d{2,4})年(\d{1,2})月[到至和-](\d{1,2})月',
        r'(\d{2,4})年(\d{1,2})[到至和-](\d{1,2})月',
    ]

    for pattern in cn_range_patterns:
        match = re.search(pattern, value)
        if match:
            year = match.group(1)
            if len(year) == 2:
                year = '20' + year
            start_month = int(match.group(2))
            end_month = int(match.group(3))
            if 1 <= start_month <= 12 and 1 <= end_month <= 12:
                months = [f"{year}{str(m).zfill(2)}" for m in range(start_month, end_month + 1)]
                return ",".join(months)

    # 处理数字日期范围 (如 202411-12)
    num_range_pattern = r'(\d{4})(\d{1,2})-(\d{1,2})'
    match = re.search(num_range_pattern, value)
    if match:
        year = match.group(1)
        start_month = int(match.group(2))
        end_month = int(match.group(3))
        if 1 <= start_month <= 12 and 1 <= end_month <= 12:
            months = [f"{year}{str(m).zfill(2)}" for m in range(start_month, end_month + 1)]
            return ",".join(months)

    return None


def _standardize_customer_name(value: str) -> str:
    """标准化客户名称"""
    value = value.replace('（', '(').replace('）', ')')
    value = value.replace('：', ':').replace('，', ',')
    value = value.replace('"', '"').replace('"', '"')
    value = value.replace('　', '')
    return value


def _standardize_product_name(value: str) -> str:
    """标准化产品名称"""
    value = value.replace('（', '(').replace('）', ')')
    value = value.replace('，', ',').replace('：', ':')
    value = value.replace('　', '')
    return value.upper()


def get_sheet_data(sheet, row: int) -> Tuple[str, str, str]:
    """获取并标准化工作表数据

    Args:
        sheet: 工作表对象
        row: 行号

    Returns:
        标准化后的数据元组 (日期, 客户名称, 产品名称)
    """
    values = tuple(
        standardize_data(str(sheet.cell(row=row, column=i).value), i)
        for i in range(1, 4)
    )
    logging.debug(f"行{row}原始数据: {values}")
    return values


def clear_sheet(sheet) -> None:
    """清空工作表数据"""
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row)


def copy_title_row(source_sheet, target_sheet) -> None:
    """复制标题行"""
    for column, cell in enumerate(source_sheet[1], start=1):
        target_sheet.cell(row=1, column=column, value=cell.value)


def init_result_sheet(workbook, sheet_name: str):
    """初始化结果工作表"""
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        clear_sheet(sheet)
    else:
        sheet = workbook.create_sheet(sheet_name)
    return sheet


def setup_logging(log_dir: str) -> str:
    """设置日志系统

    Args:
        log_dir: 日志目录路径

    Returns:
        日志文件路径
    """
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)

    # 清理历史日志
    today = datetime.now().strftime("%Y%m%d")
    try:
        for log_file_name in os.listdir(log_dir):
            if log_file_name.startswith('供应商匹配_') and log_file_name.endswith('.log'):
                file_date = log_file_name.replace('供应商匹配_', '').replace('.log', '')
                if file_date <= today:
                    old_log_path = os.path.join(log_dir, log_file_name)
                    try:
                        os.remove(old_log_path)
                        print(f"已清理历史日志: {log_file_name}")
                    except Exception as e:
                        print(f"清理日志文件失败 {log_file_name}: {str(e)}")
    except Exception as e:
        print(f"清理历史日志时出错: {str(e)}")

    # 创建新日志文件
    log_file = os.path.join(log_dir, f'供应商匹配_{today}.log')
    if os.path.exists(log_file):
        try:
            os.remove(log_file)
        except Exception as e:
            print(f"清理旧日志文件失败: {str(e)}")

    logging.basicConfig(
        filename=log_file,
        level=logging.DEBUG,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )

    return log_file


# ==================== 主窗口类 ====================

class MainWindow(QMainWindow):
    """供应商数据智能匹配系统主窗口"""

    def __init__(self):
        super().__init__()
        self.settings = QSettings('供应商数据智能匹配系统', 'DataAnalysis')
        self.log_file: Optional[str] = None

        self._init_logging()
        self._init_ui()

        logging.info("系统初始化完成")

    def _init_logging(self):
        """初始化日志系统"""
        log_dir = os.path.join(os.path.dirname(__file__), 'logs')
        self.log_file = setup_logging(log_dir)

        settings = QSettings('供应商数据智能匹配系统', 'DataAnalysis')
        enable_logging = settings.value('enable_logging', False, bool)
        logging.getLogger().disabled = not enable_logging

    def _init_ui(self):
        """初始化用户界面"""
        self.setWindowTitle("供应商数据智能匹配系统")
        self._set_window_icon()
        self.setMinimumSize(600, 400)

        # 创建中央部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # 创建标签页
        tab_widget = QTabWidget()
        main_layout.addWidget(tab_widget)

        # 添加各个标签页
        tab_widget.addTab(self._create_filter_tab(), "数据筛选")
        tab_widget.addTab(self._create_settings_tab(), "设置")

        # 添加底部按钮
        self._create_bottom_buttons(main_layout)

    def _set_window_icon(self):
        """设置窗口图标"""
        icon_path = os.path.join(os.path.dirname(__file__), 'resources', 'icon.ico')
        if getattr(sys, 'frozen', False):
            base_path = sys._MEIPASS
            icon_path = os.path.join(base_path, 'resources', 'icon.ico')

        if os.path.exists(icon_path):
            app_icon = QIcon(icon_path)
            self.setWindowIcon(app_icon)
            QApplication.setWindowIcon(app_icon)

    def _create_filter_tab(self) -> QWidget:
        """创建数据筛选标签页"""
        tab = QWidget()
        layout = QVBoxLayout(tab)

        # 添加帮助区域
        layout.addWidget(self._create_help_section())

        # 添加文件选择区域
        layout.addWidget(self._create_file_selector())

        # 添加进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)

        # 添加开始分析按钮
        analyze_button = QPushButton("开始分析")
        analyze_button.clicked.connect(self.start_analysis)
        analyze_button.setStyleSheet("""
            QPushButton { background-color: #4CAF50; color: white; padding: 8px 16px; border-radius: 4px; }
            QPushButton:hover { background-color: #45a049; }
        """)
        layout.addWidget(analyze_button)

        layout.addStretch()
        return tab

    def _create_help_section(self) -> QWidget:
        """创建帮助说明区域"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # 创建切换按钮
        help_toggle = QPushButton("显示/隐藏使用说明")
        help_toggle.setCheckable(True)
        help_toggle.setChecked(False)
        layout.addWidget(help_toggle)

        # 创建滚动区域
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll_area.setStyleSheet("""
            QScrollArea { border: none; background-color: transparent; }
            QScrollBar:vertical { border: none; background: #f0f0f0; width: 10px; margin: 0px; }
            QScrollBar::handle:vertical { background: #c0c0c0; min-height: 30px; border-radius: 5px; }
            QScrollBar::handle:vertical:hover { background: #a0a0a0; }
        """)

        # 创建帮助内容
        help_content = QLabel(self._get_help_text())
        help_content.setWordWrap(True)
        help_content.setStyleSheet("""
            QLabel {
                color: #333333;
                font-size: 13px;
                padding: 10px;
                background-color: #f8f8f8;
                border: 1px solid #e0e0e0;
                border-radius: 4px;
                margin-bottom: 10px;
            }
        """)

        scroll_area.setWidget(help_content)
        scroll_area.setVisible(False)
        layout.addWidget(scroll_area)

        # 连接信号
        help_toggle.clicked.connect(lambda checked: scroll_area.setVisible(checked))

        return widget

    def _get_help_text(self) -> str:
        """获取帮助文本内容"""
        return '''使用说明：
1. 数据准备：
   - 第一个工作表为"供应商待匹配表"，放入需要查询的数据
   - 第二个工作表为"供应商匹配原表"，放入用于匹配的数据
   - 两个工作表的前三列必须包含：日期、客户名称、产品名称
   - 请确保Excel文件中只包含这两个工作表，避免干扰分析结果

2. 数据格式要求：
   - 日期格式支持：2024-03、24年3月、3月、202411-12（会自动处理为多个月份）
     示例：2024-03、24年3月、3-4月（会自动处理为多个月份）
   - 客户名称：不区分全角半角，自动处理空格
     示例："ABC公司"与"A B C公司"会被视为相同
   - 产品名称：不区分大小写，自动处理特殊符号
     示例："Model-A"与"model a"会被视为相同

3. 操作步骤：
   1) 点击"浏览文件"选择Excel文件
   2) 确认数据格式无误后点击"开始分析"
   3) 等待分析完成，查看结果
   4) 分析完成后，结果将保存在同一Excel文件中

4. 处理结果说明：
   - 🟩绿色：表示在匹配原表中找到对应数据
   - 🟥红色：表示在匹配原表中未找到对应数据
   - 🟨黄色：表示该数据重复查询（最高优先级）
   - 🟫棕色：表示日期范围内的数据未能全部匹配成功
   - 🟪紫色：表示日期范围内的数据全部匹配成功

   颜色优先级：黄色 > 紫色/棕色 > 绿色/红色
   当一条数据符合多个条件时，将按照优先级显示颜色。
   处理逻辑:
   - 系统首先对数据进行标准化处理，统一日期格式、供应商名称和产品名称
   - 对于普通数据，直接在匹配原表中查找对应记录
   - 对于日期范围（如"3-4月"、"202411-12"），系统会检查范围内每个月份是否都能匹配
   - 当一个数据项匹配到多个供应商时，系统会为每个供应商创建单独的记录
   - 匹配结果将分别保存在"匹配到的数据"和"未找到的数据"两个工作表中

5. 常见问题：
   - 如果数据未匹配，请检查日期格式是否正确
   - 供应商名称中的空格和符号会被自动处理
   - 如果分析过程中出现错误，可以在设置中开启日志记录以便排查
   - 大量数据分析可能需要较长时间，请耐心等待'''

    def _create_file_selector(self) -> QWidget:
        """创建文件选择器"""
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 0)

        file_label = QLabel("选择数据excel")
        self.file_input = QLineEdit()
        self.file_input.setReadOnly(True)
        browse_button = QPushButton("浏览文件")
        browse_button.clicked.connect(self.browse_file)

        layout.addWidget(file_label)
        layout.addWidget(self.file_input)
        layout.addWidget(browse_button)

        return widget

    def _create_settings_tab(self) -> QWidget:
        """创建设置标签页"""
        tab = QWidget()
        layout = QVBoxLayout(tab)

        # 日志记录选项
        log_checkbox = QCheckBox("启用日志记录")
        log_checkbox.setChecked(self.settings.value('enable_logging', False, bool))
        log_checkbox.stateChanged.connect(self.toggle_logging)
        layout.addWidget(log_checkbox)

        # 日志文件位置
        log_path_label = QLabel(f"日志文件位置：{os.path.abspath(self.log_file)}")
        log_path_label.setWordWrap(True)
        layout.addWidget(log_path_label)

        layout.addStretch()
        return tab

    def _create_bottom_buttons(self, layout: QVBoxLayout):
        """创建底部按钮"""
        button_layout = QHBoxLayout()
        exit_button = QPushButton("退出")
        exit_button.clicked.connect(self.close)
        button_layout.addStretch()
        button_layout.addWidget(exit_button)
        layout.addLayout(button_layout)

    def browse_file(self):
        """浏览并选择Excel文件"""
        logging.info("开始选择文件")
        try:
            last_dir = self.settings.value('last_directory', os.path.expanduser("~/Documents"))

            file_name, _ = QFileDialog.getOpenFileName(
                self,
                "选择Excel文件",
                last_dir,
                "Excel Files (*.xlsx);;All Files (*.*)"
            )

            if file_name:
                self._validate_and_set_file(file_name)
            else:
                logging.info("未选择文件")

        except Exception as e:
            logging.error(f"选择文件时出错: {str(e)}")
            QMessageBox.critical(self, "错误", f"选择文件时出错：{str(e)}")

    def _validate_and_set_file(self, file_name: str):
        """验证并设置文件路径"""
        if not os.path.exists(file_name):
            logging.error(f"文件不存在: {file_name}")
            QMessageBox.critical(self, "错误", "所选文件不存在")
            return

        if not os.access(file_name, os.R_OK):
            logging.error(f"文件无法访问: {file_name}")
            QMessageBox.critical(self, "错误", "无法访问所选文件")
            return

        self.settings.setValue('last_directory', os.path.dirname(file_name))
        logging.info(f"选择的文件: {file_name}")
        self.file_input.setText(file_name)

    def toggle_logging(self, state: int):
        """切换日志记录状态"""
        self.settings.setValue('enable_logging', bool(state))
        logging.getLogger().disabled = not state
        status = "启用" if state else "禁用"
        logging.info(f"日志记录已{status}")

    def start_analysis(self):
        """开始数据分析"""
        logging.info("开始数据分析")
        try:
            selected_file = self.file_input.text()
            if not selected_file:
                QMessageBox.warning(self, "警告", "请选择Excel文件")
                return

            workbook = openpyxl.load_workbook(selected_file)
            logging.info(f"工作簿包含的工作表: {workbook.sheetnames}")

            if len(workbook.worksheets) < 2:
                QMessageBox.critical(self, "错误", "工作簿中缺少必要的工作表")
                return

            # 获取工作表
            sheet1 = workbook.worksheets[0]
            sheet2 = workbook.worksheets[1]
            sheet3 = init_result_sheet(workbook, "匹配到的数据")
            sheet4 = init_result_sheet(workbook, "未找到的数据")

            # 处理数据
            self.process_data(workbook, sheet1, sheet2, sheet3, sheet4)

            # 保存结果
            workbook.save(selected_file)
            logging.info("数据分析完成")
            QMessageBox.information(self, "成功", "分析完成")

        except Exception as e:
            logging.error(f"分析过程出错: {str(e)}", exc_info=True)
            QMessageBox.critical(self, "错误", f"执行分析时出错：{str(e)}")

    def process_data(self, workbook, sheet1, sheet2, sheet3, sheet4):
        """处理数据匹配逻辑

        Args:
            workbook: 工作簿对象
            sheet1: 待匹配数据表
            sheet2: 匹配原表
            sheet3: 匹配结果表
            sheet4: 未找到数据表
        """
        try:
            logging.info("开始处理数据")

            # 初始化结果表
            self._init_result_sheets(sheet1, sheet3, sheet4)

            # 预处理匹配数据
            sheet2_data = self._build_lookup_dict(sheet2)

            # 处理数据
            max_row = sheet1.max_row
            if max_row <= 1:
                raise ValueError("Sheet1中没有数据需要匹配")

            self.progress_bar.setVisible(True)
            self.progress_bar.setMaximum(max_row - 1)

            progress = QProgressDialog("努力分析中....", "取消", 0, max_row - 1, self)
            progress.setWindowModality(Qt.WindowModal)
            progress.setWindowTitle("进度")

            self._process_rows(sheet1, sheet2_data, sheet3, sheet4, progress, max_row)

            progress.setValue(max_row - 1)
            self.progress_bar.setVisible(False)

        except Exception as e:
            self.progress_bar.setVisible(False)
            logging.error(f"数据处理出错: {str(e)}", exc_info=True)
            raise

    def _init_result_sheets(self, sheet1, sheet3, sheet4):
        """初始化结果工作表"""
        copy_title_row(sheet1, sheet3)
        copy_title_row(sheet1, sheet4)
        sheet3.cell(row=1, column=4, value="供应商")
        sheet4.cell(row=1, column=4, value="供应商")

    def _build_lookup_dict(self, sheet2) -> Dict[Tuple[str, str, str], List[str]]:
        """构建查找字典

        Returns:
            键为 (日期, 客户名称, 产品名称)，值为供应商列表
        """
        sheet2_data = {}
        for row in sheet2.iter_rows(min_row=2, values_only=True):
            key = (
                standardize_data(str(row[0]), 1),  # 日期
                standardize_data(str(row[1]), 2),  # 客户公司
                standardize_data(str(row[2]), 3)   # 产品名称
            )

            if key in sheet2_data:
                sheet2_data[key].append(row[3])
            else:
                sheet2_data[key] = [row[3]]

        return sheet2_data

    def _process_rows(self, sheet1, sheet2_data, sheet3, sheet4, progress, max_row: int):
        """处理所有数据行"""
        processed_keys: Set[Tuple[str, str, str]] = set()
        date_range_map: Dict[Tuple[str, str], List[str]] = {}

        for row in range(2, max_row + 1):
            if progress.wasCanceled():
                logging.info("用户取消了操作")
                raise InterruptedError("用户取消了操作")

            self.progress_bar.setValue(row - 2)

            # 获取数据
            original_data = tuple(str(sheet1.cell(row=row, column=i).value) for i in range(1, 4))
            search_key = get_sheet_data(sheet1, row)
            logging.debug(f"处理第 {row} 行，搜索键: {search_key}")

            # 分析匹配结果
            result = self._analyze_match(search_key, sheet2_data, processed_keys, date_range_map)

            # 应用样式并保存结果
            self._apply_result(sheet1, sheet3, sheet4, row, original_data, search_key, result)

            # 记录已处理
            processed_keys.add(search_key)

    def _analyze_match(self, search_key: Tuple[str, str, str],
                       sheet2_data: Dict,
                       processed_keys: Set[Tuple[str, str, str]],
                       date_range_map: Dict[Tuple[str, str], List[str]]) -> MatchResult:
        """分析数据匹配情况"""
        result = MatchResult()

        # 检查重复
        result.is_duplicate = self._check_duplicate(search_key, processed_keys, date_range_map)

        # 处理日期范围
        if ',' in search_key[0]:
            result.is_date_range = True
            dates = search_key[0].split(',')

            # 记录日期范围
            date_range_map[search_key[1:]] = dates

            # 检查每个日期的匹配情况
            all_matches = True
            for date in dates:
                test_key = (date,) + search_key[1:]
                if test_key in sheet2_data:
                    for supplier in sheet2_data[test_key]:
                        result.matched_suppliers.append((date, supplier))
                else:
                    all_matches = False
                    logging.debug(f"未匹配的日期: {date}")

            result.is_all_match = all_matches and bool(result.matched_suppliers)

        # 检查单条数据匹配
        elif not result.is_duplicate and search_key in sheet2_data:
            result.is_match = True
            for supplier in sheet2_data[search_key]:
                result.matched_suppliers.append((search_key[0], supplier))

        return result

    def _check_duplicate(self, search_key: Tuple[str, str, str],
                         processed_keys: Set[Tuple[str, str, str]],
                         date_range_map: Dict[Tuple[str, str], List[str]]) -> bool:
        """检查是否为重复数据"""
        # 检查完全相同的键
        if search_key in processed_keys:
            logging.debug(f"检测到完全相同的重复键: {search_key}")
            return True

        # 检查单月是否在已处理的日期范围内
        if ',' not in search_key[0]:
            for range_key, months in date_range_map.items():
                if search_key[1:] == range_key and search_key[0] in months:
                    logging.debug(f"检测到日期范围内的重复: 月份 {search_key[0]} 在范围 {months} 中")
                    return True

        # 检查日期范围是否与已处理的单月数据重复
        if ',' in search_key[0]:
            dates = search_key[0].split(',')
            for date in dates:
                single_key = (date,) + search_key[1:]
                if single_key in processed_keys:
                    logging.debug(f"日期范围中的月份与已处理数据重复: {single_key}")
                    return True

        return False

    def _apply_result(self, sheet1, sheet3, sheet4, row: int, original_data: Tuple,
                      search_key: Tuple, result: MatchResult):
        """应用匹配结果（样式和数据保存）"""
        # 确定样式
        cell_style = self._determine_cell_style(result)

        # 应用样式到原始数据行
        for col in range(1, 4):
            cell = sheet1.cell(row=row, column=col)
            cell.fill = cell_style.to_pattern_fill()
            cell.font = cell_style.to_font()

        # 保存匹配结果
        self._save_match_result(sheet3, sheet4, original_data, search_key, result)

    def _determine_cell_style(self, result: MatchResult) -> CellStyle:
        """根据匹配结果确定单元格样式"""
        if result.is_duplicate:
            return CellStyles.YELLOW
        elif result.is_date_range:
            return CellStyles.PURPLE if result.is_all_match else CellStyles.BROWN
        elif result.is_match:
            return CellStyles.GREEN
        else:
            return CellStyles.RED

    def _save_match_result(self, sheet3, sheet4, original_data: Tuple,
                           search_key: Tuple, result: MatchResult):
        """保存匹配结果到相应的工作表"""
        matched_records: Set[Tuple[str, str, str]] = set()

        # 决定保存到哪个工作表
        target_sheet = sheet3 if (result.is_match or result.is_all_match) else sheet4

        if result.is_match or (result.is_date_range and result.is_all_match):
            # 保存匹配的供应商
            for _, supplier in result.matched_suppliers:
                record_key = (search_key[1], search_key[2], supplier)
                if record_key not in matched_records:
                    target_sheet.append(original_data + (supplier,))
                    matched_records.add(record_key)
        else:
            # 未匹配
            target_sheet.append(original_data + ('',))


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
